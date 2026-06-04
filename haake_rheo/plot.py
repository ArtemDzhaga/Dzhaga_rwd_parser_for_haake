#!/usr/bin/env python3
"""
Построение научных графиков HAAKE из готовой Excel-книги.

Скрипт по умолчанию читает лист `output_data`: он уже содержит первые 20 точек
и готовые пары gamma/величина для построения графиков. Excel-графики не
создаются и не меняются.
"""

from __future__ import annotations

import argparse
import math
import os
import re
import shlex
import sys
import time
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

PROJECT_DIR = Path(__file__).resolve().parent
os.environ.setdefault("MPLBACKEND", "Agg")
os.environ.setdefault("MPLCONFIGDIR", str(PROJECT_DIR / ".matplotlib-cache"))
os.environ.setdefault("XDG_CACHE_HOME", str(PROJECT_DIR / ".cache"))

plt = None
np = None
load_workbook = None


def format_elapsed(seconds: float) -> str:
    return f"{seconds:.3f} с"


class StepTimer:
    def __init__(self) -> None:
        self.started_at = time.perf_counter()
        self.last_at = self.started_at
        self.steps: list[tuple[str, float]] = []

    def mark(self, label: str) -> None:
        now = time.perf_counter()
        self.steps.append((label, now - self.last_at))
        self.last_at = now

    @property
    def total(self) -> float:
        return time.perf_counter() - self.started_at


def load_plot_dependencies() -> None:
    global load_workbook, np, plt
    if plt is not None and np is not None and load_workbook is not None:
        return
    try:
        import matplotlib.pyplot as matplotlib_pyplot
        import numpy as numpy_module
        from openpyxl import load_workbook as openpyxl_load_workbook
    except ImportError as exc:
        raise SystemExit(
            "Не хватает Python-библиотек. Установите их командой:\n"
            "python3 -m pip install matplotlib numpy openpyxl"
        ) from exc
    plt = matplotlib_pyplot
    np = numpy_module
    load_workbook = openpyxl_load_workbook


BLOCK_RE = re.compile(r"U\s*=\s*(\d+)\s*v", re.IGNORECASE)
FREQUENCY_RE = re.compile(r"([0-9]+(?:[,.][0-9]+)?)")

METRICS = {
    "g_prime": {
        "title": "G' / gamma",
        "ylabel": "G' in Pa",
        "filename": "g_prime",
    },
    "g_double_prime": {
        "title": 'G" / gamma',
        "ylabel": 'G" in Pa',
        "filename": "g_double_prime",
    },
    "eta": {
        "title": "|eta*| / gamma",
        "ylabel": "|eta*| in Pas",
        "filename": "eta",
    },
    "tan_delta": {
        "title": "tan δ / gamma",
        "ylabel": "tan δ",
        "filename": "tan_delta",
    },
}

VOLTAGE_STYLES = {
    0: {"color": "#404040", "marker": "s"},
    500: {"color": "#E53935", "marker": "o"},
    1000: {"color": "#2878D4", "marker": "^"},
    1500: {"color": "#2EAD5B", "marker": "v"},
    2000: {"color": "#8E5FD3", "marker": "D"},
    2500: {"color": "#FB8C00", "marker": "*"},
}


@dataclass(frozen=True)
class SeriesData:
    block_title: str
    label: str
    voltage: int
    frequency_hz: float
    frequency_label: str
    rows: list[dict[str, float]]


def parse_dragged_paths(raw_value: str) -> list[Path]:
    cleaned = raw_value.strip()
    if not cleaned:
        return []
    try:
        parts = shlex.split(cleaned)
    except ValueError:
        parts = [cleaned.strip("'\"")]
    return [Path(part).expanduser() for part in parts]


def prompt_line(prompt: str, default: str | None = None) -> str:
    suffix = f" [{default}]" if default else ""
    value = input(f"{prompt}{suffix}: ").strip()
    return value or (default or "")


def prompt_existing_workbook() -> Path:
    while True:
        paths = parse_dragged_paths(prompt_line("1. Укажите Excel-книгу с output_data"))
        if len(paths) != 1:
            print("Укажите ровно один .xlsx файл.")
            continue
        path = paths[0].resolve()
        if not path.is_file() or path.suffix.lower() != ".xlsx":
            print(f"Нужен существующий .xlsx файл: {path}")
            continue
        return path


def prompt_int(prompt: str, default: int, minimum: int) -> int:
    while True:
        value = prompt_line(prompt, str(default))
        try:
            result = int(value)
        except ValueError:
            print("Введите целое число.")
            continue
        if result < minimum:
            print(f"Минимальное значение: {minimum}.")
            continue
        return result


def prompt_tokens(prompt: str, default: list[str], allowed: set[str] | None = None) -> list[str]:
    default_text = " ".join(default)
    while True:
        raw_value = prompt_line(prompt, default_text)
        values = [value.strip() for value in re.split(r"[\s,;]+", raw_value) if value.strip()]
        if not values:
            return default
        if allowed is not None:
            unknown = [value for value in values if value not in allowed]
            if unknown:
                print("Неизвестные значения: " + ", ".join(unknown))
                print("Доступно: " + ", ".join(sorted(allowed)))
                continue
        return values


def prompt_yes_no(prompt: str, default: bool = False) -> bool:
    default_text = "y" if default else "n"
    value = prompt_line(prompt, default_text).strip().lower()
    return value in {"y", "yes", "д", "да"}


def interactive_args() -> argparse.Namespace:
    print("HAAKE Excel -> matplotlib-графики")
    print("Можно перетащить Excel-файл в терминал мышкой.")
    workbook = prompt_existing_workbook()
    sheet = prompt_line("2. Укажите лист с данными", "output_data")
    default_output = workbook.with_name(f"{workbook.stem}_plots")
    output_dir = Path(prompt_line("3. Укажите папку для графиков", str(default_output))).expanduser().resolve()
    points = prompt_int("4. Сколько первых точек брать", 20, 1)
    poly_order = prompt_int("5. Порядок полинома", 3, 1)
    formats = prompt_tokens("6. Форматы файлов", ["png"], {"png", "svg", "pdf"})
    metrics = prompt_tokens("7. Метрики", list(METRICS), set(METRICS))
    raw_frequencies = prompt_line("8. Частоты через пробел, пусто = все", "")
    frequencies = [value.strip() for value in re.split(r"[\s,;]+", raw_frequencies) if value.strip()] or None
    no_fit = prompt_yes_no("Отключить полиномиальные кривые?", default=False)
    dpi = prompt_int("DPI для PNG", 300, 72)
    return argparse.Namespace(
        workbook=workbook,
        sheet=sheet,
        output_dir=output_dir,
        points=points,
        poly_order=poly_order,
        curve_points=240,
        formats=formats,
        metrics=metrics,
        frequencies=frequencies,
        dpi=dpi,
        no_fit=no_fit,
    )


def parse_args(argv: list[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Построить matplotlib-графики из листа output_data Excel-книги HAAKE."
    )
    parser.add_argument("workbook", type=Path, help="Excel-книга .xlsx после запуска haake_cli.py.")
    parser.add_argument("--sheet", default="output_data", help="Лист с подготовленными данными.")
    parser.add_argument(
        "-o",
        "--output-dir",
        type=Path,
        default=Path("plots"),
        help="Папка для сохранения графиков.",
    )
    parser.add_argument(
        "--points",
        type=int,
        default=20,
        help="Сколько первых точек брать из каждого частотного подблока.",
    )
    parser.add_argument(
        "--poly-order",
        type=int,
        default=3,
        help="Порядок полинома в log10(gamma)-log10(y) координатах.",
    )
    parser.add_argument(
        "--curve-points",
        type=int,
        default=240,
        help="Количество точек для сглаживающей полиномиальной кривой.",
    )
    parser.add_argument(
        "--formats",
        nargs="+",
        default=["png"],
        choices=["png", "svg", "pdf"],
        help="Форматы файлов графиков.",
    )
    parser.add_argument(
        "--metrics",
        nargs="+",
        default=list(METRICS),
        choices=list(METRICS),
        help="Какие величины строить.",
    )
    parser.add_argument(
        "--frequencies",
        nargs="+",
        help="Ограничить частоты, например: --frequencies 0.1 0.5 1",
    )
    parser.add_argument("--dpi", type=int, default=300, help="DPI для PNG.")
    parser.add_argument(
        "--no-fit",
        action="store_true",
        help="Не рисовать полиномиальную сглаживающую линию.",
    )
    return parser.parse_args(argv)


def as_float(value) -> float | None:
    if isinstance(value, bool) or value is None:
        return None
    if isinstance(value, (int, float)):
        result = float(value)
        return result if math.isfinite(result) else None
    if isinstance(value, str):
        cleaned = value.strip().replace(",", ".")
        if not cleaned:
            return None
        try:
            result = float(cleaned)
        except ValueError:
            return None
        return result if math.isfinite(result) else None
    return None


CELL_REF_RE = re.compile(r"^=?(?:IF\()?([A-Za-z_][A-Za-z0-9_ ]*)!([A-Z]+)([0-9]+)")
DIRECT_CELL_RE = re.compile(r"^=?([A-Z]+)([0-9]+)$")


def column_index(column_letters: str) -> int:
    result = 0
    for char in column_letters.upper():
        result = result * 26 + (ord(char) - ord("A") + 1)
    return result


def cell_numeric_value(workbook, ws, row: int, column: int) -> float | None:
    value = ws.cell(row, column).value
    numeric = as_float(value)
    if numeric is not None:
        return numeric
    if not isinstance(value, str):
        return None

    formula = value.strip()
    if not formula.startswith("="):
        return None
    match = CELL_REF_RE.match(formula)
    if match:
        sheet_name, column_letters, row_text = match.groups()
        if sheet_name not in workbook.sheetnames:
            return None
        source_ws = workbook[sheet_name]
        source_value = source_ws.cell(int(row_text), column_index(column_letters)).value
        return as_float(source_value)
    direct_match = DIRECT_CELL_RE.match(formula)
    if direct_match:
        column_letters, row_text = direct_match.groups()
        source_value = ws.cell(int(row_text), column_index(column_letters)).value
        return as_float(source_value)
    return None


def parse_frequency_hz(value: str) -> float | None:
    text = str(value).lower().replace("гц", "hz").replace(",", ".")
    match = FREQUENCY_RE.search(text)
    if not match:
        return None
    return as_float(match.group(1))


def format_frequency_label(frequency_hz: float) -> str:
    text = f"{frequency_hz:g}".replace(".", ",")
    return f"{text} Hz"


def safe_frequency_name(frequency_hz: float) -> str:
    return f"{frequency_hz:g}".replace(".", "_")


def canonical_header(value) -> str | None:
    if value is None:
        return None
    text = str(value).strip().lower().replace("''", '"')
    text = text.replace("”", '"').replace("“", '"')
    if "gamma" in text:
        return "gamma"
    if "eta" in text:
        return "eta"
    if 'g"' in text:
        return "g_double_prime"
    if "g'" in text:
        return "g_prime"
    if "tau" in text:
        return "tau"
    if "t_seg" in text:
        return "t_seg"
    return None


def block_label(block_title: str, voltage: int) -> str:
    title = re.sub(r"^Частотные характеристики при\s*", "", block_title, flags=re.IGNORECASE)
    start = title.find("SCTNA")
    if start >= 0:
        title = title[start:]
    title = re.sub(r"_U\s*=\s*\d+\s*v", f"_{voltage}v", title, flags=re.IGNORECASE)
    title = re.sub(r"_30_points.*$", "", title, flags=re.IGNORECASE)
    title = title.replace(" ", "_")
    return title


def find_block_rows(ws) -> list[int]:
    rows: list[int] = []
    for row in range(1, ws.max_row + 1):
        for column in range(1, min(ws.max_column, 8) + 1):
            value = ws.cell(row, column).value
            if isinstance(value, str) and BLOCK_RE.search(value):
                rows.append(row)
                break
    return rows


def frequency_columns(ws, block_row: int) -> list[tuple[float, str, int]]:
    result: list[tuple[float, str, int]] = []
    frequency_row = block_row + 1
    for column in range(1, ws.max_column + 1):
        value = ws.cell(frequency_row, column).value
        if not isinstance(value, str):
            continue
        if "частота" not in value.lower() and "hz" not in value.lower() and "гц" not in value.lower():
            continue
        frequency = parse_frequency_hz(value)
        if frequency is None:
            continue
        result.append((frequency, format_frequency_label(frequency), column))
    return sorted(result, key=lambda item: item[0])


def header_map(ws, header_row: int, first_column: int) -> dict[str, int]:
    columns: dict[str, int] = {}
    for offset in range(0, 7):
        column = first_column + offset
        key = canonical_header(ws.cell(header_row, column).value)
        if key is not None:
            columns[key] = column
    return columns


def read_frequency_rows(
    ws,
    block_row: int,
    next_block_row: int,
    first_column: int,
    points: int,
) -> list[dict[str, float]]:
    columns = header_map(ws, block_row + 2, first_column)
    required = {"gamma", "g_prime", "g_double_prime", "eta"}
    if not required.issubset(columns):
        missing = ", ".join(sorted(required - set(columns)))
        raise RuntimeError(
            f"В блоке строки {block_row} не найдены обязательные колонки: {missing}."
        )

    rows: list[dict[str, float]] = []
    for row in range(block_row + 3, next_block_row):
        values = {name: as_float(ws.cell(row, column).value) for name, column in columns.items()}
        if all(values.get(name) is None for name in required):
            if rows:
                break
            continue
        gamma = values.get("gamma")
        g_prime = values.get("g_prime")
        g_double_prime = values.get("g_double_prime")
        eta = values.get("eta")
        if gamma is None or g_prime is None or g_double_prime is None or eta is None:
            continue
        tan_delta = g_double_prime / g_prime if g_prime != 0 else None
        if tan_delta is None:
            continue
        rows.append(
            {
                "gamma": gamma,
                "g_prime": g_prime,
                "g_double_prime": g_double_prime,
                "eta": eta,
                "tan_delta": tan_delta,
            }
        )
        if len(rows) >= points:
            break
    return rows


OUTPUT_PAIR_METRICS = {
    "g_prime": {"gamma": "gamma", "value": "g'"},
    "g_double_prime": {"gamma": "gamma", "value": 'g"'},
    "eta": {"gamma": "gamma", "value": "eta"},
}


def header_voltage(value) -> int | None:
    if value is None:
        return None
    match = re.search(r"_(\d+)v\b", str(value), re.IGNORECASE)
    return int(match.group(1)) if match else None


def output_header_metric(value) -> str | None:
    if value is None:
        return None
    text = str(value).strip().lower().replace("''", '"')
    if text.startswith("gamma_"):
        return "gamma"
    if text.startswith("g''_") or text.startswith('g"_'):
        return 'g"'
    if text.startswith("g'_"):
        return "g'"
    if "eta" in text:
        return "eta"
    if "tan" in text:
        return "tan_delta"
    return None


def output_frequency_rows(ws) -> list[tuple[int, float, str]]:
    rows: list[tuple[int, float, str]] = []
    for row in range(1, ws.max_row + 1):
        value = ws.cell(row, 1).value
        if not isinstance(value, str) or "частота" not in value.lower():
            continue
        frequency = parse_frequency_hz(value)
        if frequency is not None:
            rows.append((row, frequency, format_frequency_label(frequency)))
    return rows


def output_metric_pairs(ws, metric: str) -> list[tuple[int, int, int, str]]:
    pairs: list[tuple[int, int, int, str]] = []
    if metric in OUTPUT_PAIR_METRICS:
        wanted_gamma = OUTPUT_PAIR_METRICS[metric]["gamma"]
        wanted_value = OUTPUT_PAIR_METRICS[metric]["value"]
        for column in range(1, ws.max_column):
            left_metric = output_header_metric(ws.cell(1, column).value)
            right_metric = output_header_metric(ws.cell(1, column + 1).value)
            if left_metric != wanted_gamma or right_metric != wanted_value:
                continue
            voltage = header_voltage(ws.cell(1, column + 1).value)
            label = ws.cell(3, column + 1).value or ws.cell(3, column).value
            if voltage is not None:
                pairs.append((voltage, column, column + 1, str(label or f"{voltage}v")))
        return pairs

    if metric == "tan_delta":
        for column in range(1, ws.max_column + 1):
            if output_header_metric(ws.cell(1, column).value) != "tan_delta":
                continue
            voltage = header_voltage(ws.cell(1, column).value)
            label = ws.cell(3, column).value
            if voltage is not None:
                pairs.append((voltage, -1, column, str(label or f"{voltage}v")))
        return pairs
    return pairs


def read_output_data_series(workbook_path: Path, sheet_name: str, points: int) -> list[SeriesData]:
    workbook = load_workbook(workbook_path, data_only=False, read_only=False)
    if sheet_name not in workbook.sheetnames:
        raise RuntimeError(
            f"В книге нет листа {sheet_name!r}. Доступные листы: {', '.join(workbook.sheetnames)}"
        )
    ws = workbook[sheet_name]
    frequency_rows = output_frequency_rows(ws)
    if not frequency_rows:
        raise RuntimeError(f"На листе {sheet_name!r} не найдены строки вида 'Частота ... Гц'.")

    by_key: dict[tuple[float, int, str], SeriesData] = {}
    for metric in METRICS:
        for voltage, gamma_column, value_column, label in output_metric_pairs(ws, metric):
            for frequency_row, frequency, frequency_label in frequency_rows:
                rows = by_key.setdefault(
                    (frequency, voltage, label),
                    SeriesData(
                        block_title=sheet_name,
                        label=label,
                        voltage=voltage,
                        frequency_hz=frequency,
                        frequency_label=frequency_label,
                        rows=[],
                    ),
                ).rows
                for offset in range(points):
                    row = frequency_row + 1 + offset
                    if row > ws.max_row:
                        break
                    if metric == "tan_delta":
                        gamma_column_for_voltage = next(
                            (
                                pair_gamma_column
                                for pair_voltage, pair_gamma_column, _, _ in output_metric_pairs(ws, "g_prime")
                                if pair_voltage == voltage
                            ),
                            None,
                        )
                        if gamma_column_for_voltage is None:
                            continue
                        gamma = cell_numeric_value(workbook, ws, row, gamma_column_for_voltage)
                    else:
                        gamma = cell_numeric_value(workbook, ws, row, gamma_column)
                    value = cell_numeric_value(workbook, ws, row, value_column)
                    if gamma is None or value is None:
                        continue
                    while len(rows) <= offset:
                        rows.append({})
                    rows[offset]["gamma"] = gamma
                    rows[offset][metric] = value

    series: list[SeriesData] = []
    for item in by_key.values():
        completed_rows = []
        for row in item.rows[:points]:
            gamma = row.get("gamma")
            g_prime = row.get("g_prime")
            g_double_prime = row.get("g_double_prime")
            eta = row.get("eta")
            tan_delta = row.get("tan_delta")
            if tan_delta is None and g_prime not in (None, 0) and g_double_prime is not None:
                tan_delta = g_double_prime / g_prime
            if gamma is None:
                continue
            completed = {"gamma": gamma}
            if g_prime is not None:
                completed["g_prime"] = g_prime
            if g_double_prime is not None:
                completed["g_double_prime"] = g_double_prime
            if eta is not None:
                completed["eta"] = eta
            if tan_delta is not None:
                completed["tan_delta"] = tan_delta
            completed_rows.append(completed)
        if completed_rows:
            series.append(
                SeriesData(
                    block_title=item.block_title,
                    label=item.label,
                    voltage=item.voltage,
                    frequency_hz=item.frequency_hz,
                    frequency_label=item.frequency_label,
                    rows=completed_rows,
                )
            )
    return sorted(series, key=lambda item: (item.frequency_hz, item.voltage, item.label))


def read_input_data_series(workbook_path: Path, sheet_name: str, points: int) -> list[SeriesData]:
    workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    if sheet_name not in workbook.sheetnames:
        raise RuntimeError(
            f"В книге нет листа {sheet_name!r}. Доступные листы: {', '.join(workbook.sheetnames)}"
        )
    ws = workbook[sheet_name]
    block_rows = find_block_rows(ws)
    if not block_rows:
        raise RuntimeError(f"На листе {sheet_name!r} не найдены блоки с U=...v.")

    series: list[SeriesData] = []
    row_limits = block_rows[1:] + [ws.max_row + 1]
    for block_row, next_block_row in zip(block_rows, row_limits):
        block_title = str(ws.cell(block_row, 2).value or "")
        voltage_match = BLOCK_RE.search(block_title)
        if voltage_match is None:
            continue
        voltage = int(voltage_match.group(1))
        label = block_label(block_title, voltage)
        for frequency, frequency_label, first_column in frequency_columns(ws, block_row):
            rows = read_frequency_rows(ws, block_row, next_block_row, first_column, points)
            if not rows:
                continue
            series.append(
                SeriesData(
                    block_title=block_title,
                    label=label,
                    voltage=voltage,
                    frequency_hz=frequency,
                    frequency_label=frequency_label,
                    rows=rows,
                )
            )
    return sorted(series, key=lambda item: (item.frequency_hz, item.voltage, item.label))


def read_series(workbook_path: Path, sheet_name: str, points: int) -> list[SeriesData]:
    if sheet_name.lower() == "output_data":
        return read_output_data_series(workbook_path, sheet_name, points)
    return read_input_data_series(workbook_path, sheet_name, points)


def positive_metric_points(series: SeriesData, metric: str) -> tuple[np.ndarray, np.ndarray]:
    pairs = [
        (row["gamma"], row[metric])
        for row in series.rows
        if row.get("gamma") is not None
        and row.get(metric) is not None
        and row["gamma"] > 0
        and row[metric] > 0
    ]
    pairs.sort(key=lambda item: item[0])
    if not pairs:
        return np.array([]), np.array([])
    xs, ys = zip(*pairs)
    return np.array(xs, dtype=float), np.array(ys, dtype=float)


def fitted_curve(
    xs: np.ndarray,
    ys: np.ndarray,
    order: int,
    curve_points: int,
) -> tuple[np.ndarray, np.ndarray]:
    if len(xs) < 3:
        return np.array([]), np.array([])
    effective_order = min(max(order, 1), len(xs) - 1)
    log_xs = np.log10(xs)
    log_ys = np.log10(ys)
    coefficients = np.polyfit(log_xs, log_ys, effective_order)
    curve_log_xs = np.linspace(float(log_xs.min()), float(log_xs.max()), curve_points)
    curve_log_ys = np.polyval(coefficients, curve_log_xs)
    return np.power(10.0, curve_log_xs), np.power(10.0, curve_log_ys)


def grouped_by_frequency(series: Iterable[SeriesData]) -> dict[float, list[SeriesData]]:
    groups: dict[float, list[SeriesData]] = {}
    for item in series:
        groups.setdefault(item.frequency_hz, []).append(item)
    return groups


def selected_frequencies(values: list[str] | None) -> set[float] | None:
    if not values:
        return None
    selected: set[float] = set()
    for value in values:
        frequency = parse_frequency_hz(value)
        if frequency is None:
            raise RuntimeError(f"Не удалось разобрать частоту: {value!r}")
        selected.add(round(frequency, 10))
    return selected


def style_for_voltage(voltage: int) -> dict[str, str]:
    return VOLTAGE_STYLES.get(voltage, {"color": "#111111", "marker": "o"})


def plot_metric(
    frequency_hz: float,
    frequency_series: list[SeriesData],
    metric: str,
    output_dir: Path,
    formats: list[str],
    dpi: int,
    poly_order: int,
    curve_points: int,
    draw_fit: bool,
) -> list[Path]:
    metric_info = METRICS[metric]
    fig, ax = plt.subplots(figsize=(7.2, 8.6), dpi=dpi)
    plotted = 0

    for item in frequency_series:
        xs, ys = positive_metric_points(item, metric)
        if len(xs) == 0:
            continue
        style = style_for_voltage(item.voltage)
        ax.scatter(
            xs,
            ys,
            s=34,
            marker=style["marker"],
            color=style["color"],
            edgecolor=style["color"],
            linewidths=0.8,
            label=item.label,
            zorder=3,
        )
        if draw_fit:
            curve_xs, curve_ys = fitted_curve(xs, ys, poly_order, curve_points)
            if len(curve_xs) > 0:
                ax.plot(
                    curve_xs,
                    curve_ys,
                    color=style["color"],
                    linewidth=1.25,
                    label="_nolegend_",
                    zorder=2,
                )
        plotted += 1

    if plotted == 0:
        plt.close(fig)
        return []

    frequency_label = format_frequency_label(frequency_hz)
    ax.set_xscale("log")
    ax.set_yscale("log")
    ax.set_box_aspect(1)
    ax.set_xlabel(f"Gamma {frequency_label.replace(' ', '')}", fontsize=13, fontweight="bold")
    ax.set_ylabel(f"{metric_info['ylabel']} {frequency_label.replace(' ', '')}", fontsize=13, fontweight="bold")
    ax.set_title(f"{metric_info['title']} | {frequency_label}", fontsize=12, fontweight="bold")
    ax.grid(True, which="major", linestyle=":", linewidth=1.0, color="#B8B8B8", alpha=0.9)
    ax.grid(True, which="minor", linestyle=":", linewidth=0.7, color="#D0D0D0", alpha=0.75)
    ax.tick_params(axis="both", which="major", direction="in", top=True, right=True, length=7, width=1.1)
    ax.tick_params(axis="both", which="minor", direction="in", top=True, right=True, length=4, width=0.9)
    for spine in ax.spines.values():
        spine.set_linewidth(1.1)
    legend = ax.legend(
        loc="upper center",
        bbox_to_anchor=(0.5, -0.2),
        ncol=1,
        frameon=True,
        fancybox=False,
        framealpha=1.0,
        edgecolor="#222222",
        fontsize=10,
        markerscale=1.15,
        columnspacing=1.6,
        handletextpad=0.8,
    )
    legend.get_frame().set_linewidth(1.0)
    fig.tight_layout(rect=(0, 0.24, 1, 1))

    saved: list[Path] = []
    output_dir.mkdir(parents=True, exist_ok=True)
    stem = f"{safe_frequency_name(frequency_hz)}Hz_{metric_info['filename']}"
    for file_format in formats:
        path = output_dir / f"{stem}.{file_format}"
        fig.savefig(path, dpi=dpi, bbox_inches="tight")
        saved.append(path)
    plt.close(fig)
    return saved


def main(argv: list[str] | None = None) -> int:
    timer = StepTimer()
    argv = sys.argv[1:] if argv is None else argv
    args = interactive_args() if not argv else parse_args(argv)
    if not args.workbook.exists():
        raise RuntimeError(f"Excel-книга не найдена: {args.workbook}")
    if args.points <= 0:
        raise RuntimeError("--points должен быть больше 0.")
    if args.poly_order < 1:
        raise RuntimeError("--poly-order должен быть больше 0.")
    if args.curve_points < 20:
        raise RuntimeError("--curve-points должен быть не меньше 20.")
    timer.mark("Валидация аргументов")

    load_plot_dependencies()
    timer.mark("Загрузка библиотек")
    series = read_series(args.workbook, args.sheet, args.points)
    timer.mark("Чтение данных Excel")
    frequency_filter = selected_frequencies(args.frequencies)
    groups = grouped_by_frequency(series)
    saved: list[Path] = []
    for frequency_hz, frequency_series in sorted(groups.items()):
        if frequency_filter is not None and round(frequency_hz, 10) not in frequency_filter:
            continue
        for metric in args.metrics:
            saved.extend(
                plot_metric(
                    frequency_hz,
                    frequency_series,
                    metric,
                    args.output_dir,
                    args.formats,
                    args.dpi,
                    args.poly_order,
                    args.curve_points,
                    not args.no_fit,
                )
            )

    if not saved:
        raise RuntimeError("Не построено ни одного графика: проверьте лист, частоты и данные.")

    print(f"Найдено серий данных: {len(series)}")
    print(f"Создано файлов графиков: {len(saved)}")
    print(f"Папка: {args.output_dir.resolve()}")
    timer.mark("Построение и сохранение графиков")
    print("Время выполнения:")
    for label, seconds in timer.steps:
        print(f"  - {label}: {format_elapsed(seconds)}")
    print(f"  - Итого: {format_elapsed(timer.total)}")
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except RuntimeError as exc:
        print(f"Ошибка: {exc}", file=sys.stderr)
        raise SystemExit(1)
