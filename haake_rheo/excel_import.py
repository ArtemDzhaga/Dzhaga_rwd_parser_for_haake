#!/usr/bin/env python3
"""
Локальный CLI для переноса данных HAAKE RheoWin .rwd в Excel-шаблон.

Порядок аргументов:
1. Excel-шаблон .xlsx.
2. ASCII-экспорт RheoWin .txt/.asc/.csv со структурой колонок.
3. Один или несколько .rwd файлов либо папок с .rwd.

ASCII используется как описание структуры: скрипт берет из него официальные
названия колонок, количество строк в частотном сегменте и список частотных
сегментов. Значения переносятся из переданных .rwd файлов.
"""

from __future__ import annotations

import argparse
import csv
import math
import re
import shlex
import struct
import sys
import time
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


NUMERIC_RECORD_MARKER = bytes.fromhex("da00da00feff")
TEXT_RE = re.compile(rb"[\x09\x0a\x0d\x20-\x7e]{4,}")
ASCII_SUFFIXES = {".txt", ".asc", ".csv"}


def format_elapsed(seconds: float) -> str:
    return f"{seconds:.3f} с"


class StepTimer:
    def __init__(self) -> None:
        self.started_at = time.perf_counter()
        self.last_at = self.started_at
        self.steps: list[tuple[str, float]] = []

    def mark(self, label: str) -> None:
        now = time.perf_counter()
        self.steps.append((label, now - self.last_at))
        self.last_at = now

    @property
    def total(self) -> float:
        return time.perf_counter() - self.started_at


@dataclass
class NumericValue:
    value: float
    raw_hex_le: str
    offset: int


@dataclass
class NumericBlock:
    offset: int
    data_offset: int
    code_1: int
    code_2: int
    dtype: int
    values: list[NumericValue]
    extended_values: list[NumericValue]

    @property
    def key(self) -> tuple[int, int, int]:
        return self.code_1, self.code_2, self.dtype


@dataclass
class RwdRecord:
    path: Path
    blob: bytes
    size_bytes: int
    format_magic: str
    rheowin_version: str
    date: str
    time: str
    device: str
    instrument_model: str
    geometry: str
    temperature_controller: str
    gap: str
    job_file: str
    mode: str
    serial_numbers: str
    driver_versions: str
    firmware_versions: str
    strings: list[tuple[int, str]]
    numeric_blocks: list[NumericBlock]


@dataclass
class AsciiRecord:
    path: Path
    source_rwd_path: str
    operator: str
    date_time_version: str
    headers: list[str]
    rows: list[list[str | float]]

    @property
    def row_labels(self) -> list[str]:
        return [str(row[0]) for row in self.rows]


@dataclass
class ChannelBinding:
    header: str
    raw_offset: int
    segment_offsets: tuple[int, ...] = ()
    block_key: tuple[int, int, int] | None = None
    block_occurrence: int = 0
    segment_stride: int = 0


@dataclass(frozen=True)
class AsciiSegment:
    label: str
    start_index: int
    row_count: int
    frequency_sort_key: float
    frequency_label: str


@dataclass(frozen=True)
class ExperimentIdentity:
    block_name: str
    voltage: str
    frequency: str
    frequency_sort_key: float | None
    is_batch: bool = False


DISPLAY_COLUMNS = [
    ("t_seg in s", "t_seg in s"),
    ("Tau in Pa", "Tau in Pa"),
    ("G' in Pa", "G' in Pa"),
    ('G" in Pa', 'G" in Pa'),
    ("|Eta*| in Pas", "|eta*| in Pas"),
    ("Gamma in -", "gamma"),
]


def extract_ascii_strings(blob: bytes) -> list[tuple[int, str]]:
    rows: list[tuple[int, str]] = []
    for match in TEXT_RE.finditer(blob):
        text = match.group().decode("latin1", errors="replace").strip()
        if text:
            rows.append((match.start(), text))
    return rows


def extract_numeric_blocks(blob: bytes) -> list[NumericBlock]:
    blocks: list[NumericBlock] = []
    start = 0
    while True:
        offset = blob.find(NUMERIC_RECORD_MARKER, start)
        if offset < 0:
            break
        start = offset + 1

        if offset + 60 > len(blob):
            continue

        code_1 = struct.unpack_from("<H", blob, offset + 6)[0]
        code_2 = struct.unpack_from("<H", blob, offset + 8)[0]
        dtype = struct.unpack_from("<I", blob, offset + 14)[0]
        count = struct.unpack_from("<I", blob, offset + 18)[0]
        if count <= 0 or count > 10000:
            count16 = struct.unpack_from("<H", blob, offset + 18)[0]
            if 0 < count16 <= 10000:
                count = count16
        if dtype != 4 or count <= 0 or count > 10000:
            continue

        data_offset = offset + 60
        data_end = data_offset + count * 4
        if data_end > len(blob):
            continue

        # Some RheoWin job-manager files declare half of the physical array
        # length in the record header. Keep one extra declared-length chunk
        # because the second half can contain the remaining measurement points.
        extended_count = min(count * 2, (len(blob) - data_offset) // 4)
        extended_values: list[NumericValue] = []
        for index in range(extended_count):
            value_offset = data_offset + index * 4
            raw_bytes = blob[value_offset : value_offset + 4]
            extended_values.append(
                NumericValue(
                    value=struct.unpack("<f", raw_bytes)[0],
                    raw_hex_le=raw_bytes.hex(),
                    offset=value_offset,
                )
            )
        blocks.append(
            NumericBlock(
                offset=offset,
                data_offset=data_offset,
                code_1=code_1,
                code_2=code_2,
                dtype=dtype,
                values=extended_values[:count],
                extended_values=extended_values,
            )
        )
    return blocks


def unique_join(values: list[str]) -> str:
    seen: set[str] = set()
    result: list[str] = []
    for value in values:
        cleaned = value.strip()
        if cleaned and cleaned not in seen:
            seen.add(cleaned)
            result.append(cleaned)
    return " | ".join(result)


def first_match(pattern: str, text: str, flags: int = 0) -> str:
    match = re.search(pattern, text, flags)
    return match.group(1).strip() if match else ""


def all_matches(pattern: str, text: str, flags: int = 0) -> list[str]:
    return [match.strip() for match in re.findall(pattern, text, flags) if match.strip()]


def parse_pipe_metadata(text: str, key: str) -> list[str]:
    return all_matches(rf"{re.escape(key)}:([^|$\n\r]+)", text)


def detect_device(lines: list[str], text: str) -> str:
    candidates: list[str] = []
    for line in lines:
        if "\\" in line or ".pdf" in line.lower() or ".rwj" in line.lower():
            continue
        if re.search(r"\bRheoStress\b|\bRheoWin\b|\bRS1500?\b", line, re.I):
            if not line.startswith("***") and len(line) < 120:
                candidates.append(line)
    candidates.extend(all_matches(r"(RheoStress\s+RS\d+)", text, re.I))
    candidates.extend(all_matches(r"\b(RS\d{2,4})\b", text))
    return unique_join(candidates[:8])


def detect_geometry(lines: list[str], text: str) -> str:
    candidates: list[str] = []
    for line in lines:
        if re.search(r"\b(PP|CP|DG|Z|C)\d+\b|plate|cone|sensor", line, re.I) and len(line) < 120:
            candidates.append(line)
    candidates.extend(all_matches(r"\b(PP\d+\s*[A-Z0-9 ]*)", text))
    return unique_join(candidates[:8])


def extract_rwd_record(path: Path) -> RwdRecord:
    blob = path.read_bytes()
    strings = extract_ascii_strings(blob)
    lines = [text for _, text in strings]
    full_text = "\n".join(lines)
    date_time = re.search(r"Date/Time:\s*([0-9.]+)\s*/\s*([0-9:]+)", full_text)

    firmware_versions = parse_pipe_metadata(full_text, "Firmware version 1")
    firmware_versions.extend(parse_pipe_metadata(full_text, "Firmware version 3"))
    return RwdRecord(
        path=path,
        blob=blob,
        size_bytes=len(blob),
        format_magic="OOP_HADES" if b"OOP_HADES" in blob else blob[:16].hex(" "),
        rheowin_version=first_match(r"HAAKE RheoWin\s+([0-9.]+)", full_text),
        date=date_time.group(1) if date_time else first_match(r"\b([0-3]\d\.[01]\d\.\d{4})\b", full_text),
        time=date_time.group(2) if date_time else first_match(r"\b([0-2]\d:[0-5]\d:[0-5]\d)\b", full_text),
        device=detect_device(lines, full_text),
        instrument_model=unique_join(all_matches(r"\b(RheoStress\s+RS\d+|RS1500?|DC50)\b", full_text, re.I)),
        geometry=detect_geometry(lines, full_text),
        temperature_controller=unique_join(all_matches(r"\b(DC50\b[^\n\r]*)", full_text)),
        gap=first_match(r"Gap:\s*([^\n\r]+)", full_text),
        job_file=unique_join(all_matches(r"([A-Z]:\\[^\n\r]+?\.rwj)", full_text, re.I)),
        mode=first_match(r"ElmStros::Execute:\s*Mode:\s*([^\n\r]+)", full_text),
        serial_numbers=unique_join(parse_pipe_metadata(full_text, "Serial number")),
        driver_versions=unique_join(parse_pipe_metadata(full_text, "Driver version")),
        firmware_versions=unique_join(firmware_versions),
        strings=strings,
        numeric_blocks=extract_numeric_blocks(blob),
    )


def read_ascii_text(path: Path) -> str:
    for encoding in ("utf-8-sig", "cp1251", "cp1252", "latin1"):
        try:
            return path.read_text(encoding=encoding)
        except UnicodeError:
            continue
    raise RuntimeError(f"Не удалось прочитать ASCII-файл: {path}")


def parse_ascii_number(value: str) -> str | float:
    stripped = value.strip()
    if not stripped:
        return ""
    try:
        return float(stripped.replace(",", "."))
    except ValueError:
        return stripped


def extract_ascii_record(path: Path) -> AsciiRecord:
    lines = read_ascii_text(path).splitlines()
    header_index = next(
        (index for index, line in enumerate(lines) if line.startswith(";") and line.count(";") >= 2),
        None,
    )
    if header_index is None:
        raise RuntimeError(f"В ASCII-файле не найдена строка заголовков RheoWin: {path}")

    headers = next(csv.reader([lines[header_index]], delimiter=";"))
    headers = ["segment_point" if index == 0 and not header else header.strip() for index, header in enumerate(headers)]
    while headers and not headers[-1]:
        headers.pop()

    rows: list[list[str | float]] = []
    for line in lines[header_index + 1 :]:
        if not line.strip():
            continue
        cells = next(csv.reader([line], delimiter=";"))
        while cells and not cells[-1]:
            cells.pop()
        if len(cells) == len(headers):
            rows.append([cells[0].strip()] + [parse_ascii_number(cell) for cell in cells[1:]])

    if not rows:
        raise RuntimeError(f"В ASCII-файле нет строк измерений: {path}")

    metadata = [line.strip() for line in lines[:header_index] if line.strip()]
    return AsciiRecord(
        path=path,
        source_rwd_path=metadata[0] if metadata else "",
        operator=metadata[1] if len(metadata) > 1 else "",
        date_time_version=metadata[2] if len(metadata) > 2 else "",
        headers=headers,
        rows=rows,
    )


def value_matches(exported: float, actual: float) -> bool:
    if not math.isfinite(actual):
        return False
    tolerance = max(1e-7, abs(exported) * 0.001)
    return abs(exported - actual) <= tolerance


def ascii_column_values(ascii_record: AsciiRecord, column_index: int) -> list[float | None]:
    values: list[float | None] = []
    for row in ascii_record.rows:
        value = row[column_index]
        values.append(float(value) if isinstance(value, float) else None)
    return values


def nth_numeric_block(
    record: RwdRecord,
    key: tuple[int, int, int],
    occurrence: int,
) -> NumericBlock | None:
    current_occurrence = 0
    for block in record.numeric_blocks:
        if block.key != key:
            continue
        if current_occurrence == occurrence:
            return block
        current_occurrence += 1
    return None


def required_ascii_columns(ascii_record: AsciiRecord) -> list[tuple[int, str, list[float | None]]]:
    required = {normalize_header(header) for header, _ in DISPLAY_COLUMNS}
    required.add(normalize_header("f in Hz"))
    return [
        (column_index, header, ascii_column_values(ascii_record, column_index))
        for column_index, header in enumerate(ascii_record.headers[1:], start=1)
        if normalize_header(header) in required
    ]


def frequency_column_index(ascii_record: AsciiRecord) -> int:
    for index, header in enumerate(ascii_record.headers):
        if normalize_header(header) == normalize_header("f in Hz"):
            return index
    raise RuntimeError("В ASCII структуры отсутствует обязательная колонка 'f in Hz'.")


def format_frequency_token(frequency: float) -> str:
    if math.isclose(frequency, round(frequency), rel_tol=1e-9, abs_tol=1e-9):
        return f"{int(round(frequency))}hz"
    if 0 < frequency < 1:
        decimal_digit = int(round(frequency * 10))
        if math.isclose(frequency, decimal_digit / 10, rel_tol=1e-9, abs_tol=1e-9):
            return f"0{decimal_digit}hz"
    text = f"{frequency:g}".replace(".", "")
    return f"{text}hz"


def ascii_segments(ascii_record: AsciiRecord) -> list[AsciiSegment]:
    frequency_index = frequency_column_index(ascii_record)
    segments: list[AsciiSegment] = []
    current_label = ""
    current_start = 0

    def add_segment(label: str, start_index: int, end_index: int) -> None:
        values = [
            row[frequency_index]
            for row in ascii_record.rows[start_index:end_index]
            if isinstance(row[frequency_index], float)
        ]
        if not values:
            raise RuntimeError(f"В ASCII-сегменте {label!r} не найдена числовая частота.")
        frequency = values[0]
        if not all(value_matches(frequency, value) for value in values):
            raise RuntimeError(
                f"ASCII-сегмент {label!r} содержит разные значения 'f in Hz'."
            )
        segments.append(
            AsciiSegment(
                label=label,
                start_index=start_index,
                row_count=end_index - start_index,
                frequency_sort_key=frequency,
                frequency_label=format_frequency_token(frequency),
            )
        )

    for index, label in enumerate(ascii_record.row_labels):
        segment_label = label.split("|", 1)[0] if "|" in label else "1"
        if index == 0:
            current_label = segment_label
            current_start = 0
            continue
        if segment_label != current_label:
            add_segment(current_label, current_start, index)
            current_label = segment_label
            current_start = index
    add_segment(current_label or "1", current_start, len(ascii_record.rows))

    return segments


def resolve_values(record: RwdRecord, binding: ChannelBinding, row_count: int) -> list[NumericValue]:
    return resolve_values_slice(record, binding, 0, row_count, 0 if binding.segment_offsets else None)


def resolve_values_slice(
    record: RwdRecord,
    binding: ChannelBinding,
    start_index: int,
    row_count: int,
    segment_index: int | None = None,
) -> list[NumericValue]:
    if binding.segment_offsets:
        if segment_index is None:
            raise RuntimeError(
                f"Канал {binding.header}: для пакетного ASCII нужен номер сегмента."
            )
        if segment_index >= len(binding.segment_offsets):
            raise RuntimeError(
                f"Канал {binding.header}: сегмент {segment_index + 1} отсутствует в сопоставлении."
            )
        start_offset = binding.segment_offsets[segment_index]
    else:
        start_offset = binding.raw_offset + start_index * 4
    end_offset = start_offset + row_count * 4
    if end_offset > len(record.blob):
        raise RuntimeError(
            f"Файл {record.path.name}: отсутствует полный массив канала {binding.header} "
            f"по доказанному смещению {hex(binding.raw_offset)}."
        )
    values: list[NumericValue] = []
    for offset in range(start_offset, end_offset, 4):
        raw_bytes = record.blob[offset : offset + 4]
        values.append(
            NumericValue(
                value=struct.unpack("<f", raw_bytes)[0],
                raw_hex_le=raw_bytes.hex(),
                offset=offset,
            )
        )
    return values


def normalize_header(header: str) -> str:
    return re.sub(r"\s+", " ", header.strip()).lower().replace("''", '"')


STRUCTURAL_CHANNELS = {
    normalize_header(header): binding
    for header, binding in {
        "t_seg in s": ((7, 5, 4), 0),
        "Tau in Pa": ((2, 0, 4), 0),
        "G' in Pa": ((21, 1, 4), 0),
        'G" in Pa': ((21, 2, 4), 0),
        "|Eta*| in Pas": ((17, 3, 4), 0),
        "Gamma in -": ((16, 0, 4), 0),
        "f in Hz": ((35, 0, 4), 0),
    }.items()
}


def build_structural_bindings(ascii_record: AsciiRecord, records: list[RwdRecord]) -> list[ChannelBinding]:
    segments = ascii_segments(ascii_record)
    row_counts = {segment.row_count for segment in segments}
    if len(row_counts) != 1:
        raise RuntimeError(
            "ASCII содержит частотные сегменты разной длины. "
            "Для такой структуры нужен отдельный режим разметки."
        )
    row_count = next(iter(row_counts))
    segment_stride = row_count * 4 + 18
    has_multiple_segments = len(segments) > 1

    bindings: list[ChannelBinding] = []
    missing_headers: list[str] = []
    for _, header, _ in required_ascii_columns(ascii_record):
        channel = STRUCTURAL_CHANNELS.get(normalize_header(header))
        if channel is None:
            missing_headers.append(header)
            continue
        block_key, occurrence = channel
        bindings.append(
            ChannelBinding(
                header=header,
                raw_offset=0,
                segment_offsets=tuple(range(len(segments))) if has_multiple_segments else (),
                block_key=block_key,
                block_occurrence=occurrence,
                segment_stride=segment_stride if has_multiple_segments else 0,
            )
        )
    if missing_headers:
        raise RuntimeError(
            "В скрипте нет структурного соответствия для колонок ASCII: "
            + ", ".join(missing_headers)
        )

    for record in records:
        missing_blocks = [
            f"{binding.header} -> {binding.block_key}"
            for binding in bindings
            if binding.block_key is not None
            and nth_numeric_block(record, binding.block_key, binding.block_occurrence) is None
        ]
        if missing_blocks:
            raise RuntimeError(
                f"Файл {record.path.name}: не найдены ожидаемые бинарные каналы: "
                + ", ".join(missing_blocks)
                + ". Для этого файла нужен другой ASCII/режим структуры."
            )
    return bindings


def display_bindings(bindings: list[ChannelBinding]) -> list[tuple[str, ChannelBinding]]:
    by_header = {normalize_header(binding.header): binding for binding in bindings}
    result: list[tuple[str, ChannelBinding]] = []
    missing: list[str] = []
    for ascii_header, display_header in DISPLAY_COLUMNS:
        binding = by_header.get(normalize_header(ascii_header))
        if binding is None:
            missing.append(ascii_header)
        else:
            result.append((display_header, binding))
    if missing:
        raise RuntimeError(
            "В ASCII структуры отсутствуют обязательные рабочие колонки: "
            + ", ".join(missing)
        )
    return result


def adapt_bindings_to_record(
    structural_bindings: list[ChannelBinding],
    record: RwdRecord,
) -> list[ChannelBinding]:
    adapted: list[ChannelBinding] = []
    for binding in structural_bindings:
        if binding.block_key is None:
            adapted.append(binding)
            continue
        target_block = nth_numeric_block(record, binding.block_key, binding.block_occurrence)
        if target_block is None:
            raise RuntimeError(
                f"Файл {record.path.name}: не найден бинарный блок {binding.block_key} "
                f"для канала {binding.header!r}."
            )
        if binding.segment_offsets:
            offsets = tuple(
                target_block.data_offset + segment_index * binding.segment_stride
                for segment_index in range(len(binding.segment_offsets))
            )
        else:
            offsets = ()
        adapted.append(
            ChannelBinding(
                header=binding.header,
                raw_offset=target_block.data_offset,
                segment_offsets=offsets,
                block_key=binding.block_key,
                block_occurrence=binding.block_occurrence,
                segment_stride=binding.segment_stride,
            )
        )
    return adapted


def matching_frequency_prefix(values: list[NumericValue], frequency: float) -> int:
    count = 0
    for item in values:
        if not math.isfinite(item.value) or not value_matches(frequency, item.value):
            break
        count += 1
    return count


def minimum_segment_points(row_count: int) -> int:
    return max(3, row_count // 2)


def available_ascii_segment_indexes(
    record: RwdRecord,
    bindings: list[ChannelBinding],
    segments: list[AsciiSegment],
) -> list[tuple[int, int, int]]:
    frequency_binding = next(
        (
            binding
            for binding in bindings
            if normalize_header(binding.header) == normalize_header("f in Hz")
        ),
        None,
    )
    if frequency_binding is None:
        raise RuntimeError("В ASCII структуры отсутствует обязательная колонка 'f in Hz'.")
    if not frequency_binding.segment_offsets:
        return [(0, 0, segments[0].row_count)]

    result: list[tuple[int, int, int]] = []
    used: set[int] = set()
    for segment_offset_index in range(len(frequency_binding.segment_offsets)):
        try:
            values = resolve_values_slice(
                record,
                frequency_binding,
                0,
                segments[0].row_count,
                segment_offset_index,
            )
        except RuntimeError:
            continue
        for ascii_index, segment in enumerate(segments):
            if ascii_index in used:
                continue
            matched_points = matching_frequency_prefix(values, segment.frequency_sort_key)
            if matched_points >= minimum_segment_points(segment.row_count):
                result.append((ascii_index, segment_offset_index, min(matched_points, segment.row_count)))
                used.add(ascii_index)
                break
    if not result:
        raise RuntimeError(
            f"Файл {record.path.name}: не удалось сопоставить частотные сегменты с ASCII."
        )
    return result


def extract_signed_rows(
    record: RwdRecord,
    row_count: int,
    selected_bindings: list[tuple[str, ChannelBinding]],
    start_index: int = 0,
    segment_index: int | None = None,
) -> list[list[str | float]]:
    value_sets = [
        resolve_values_slice(record, binding, start_index, row_count, segment_index)
        for _, binding in selected_bindings
    ]
    rows: list[list[str | float]] = []
    for index in range(row_count):
        rows.append(
            [f"1|{index + 1}"]
            + [excel_value(values[index].value) for values in value_sets]
        )
    return rows


def excel_value(value: float) -> float | str:
    if math.isnan(value):
        return "NaN"
    if math.isinf(value):
        return "+Inf" if value > 0 else "-Inf"
    return value


def collect_rwd_files(paths: list[Path], recursive: bool) -> list[Path]:
    files: list[Path] = []
    for path in paths:
        expanded = path.expanduser()
        if expanded.is_dir():
            files.extend(expanded.glob("**/*.rwd" if recursive else "*.rwd"))
        elif expanded.is_file() and expanded.suffix.lower() == ".rwd":
            files.append(expanded)
        else:
            visible_path = str(path).replace("\n", "\\n")
            raise RuntimeError(
                f"Не найден переданный .rwd файл или папка: {visible_path!r}. "
                "Проверьте, что путь не содержит перенос строки внутри кавычек."
            )
    return sorted(set(file.resolve() for file in files), key=lambda path: path.name.lower())


def parse_experiment_identity(path: Path) -> ExperimentIdentity:
    match = re.match(
        r"^(?P<prefix>.+?)_freq=(?P<frequency>.+?)_(?P<voltage>U=[^_]+)(?:_.*)?$",
        path.stem,
        re.I,
    )
    if not match:
        raise RuntimeError(
            f"Не удалось разобрать имя файла {path.name!r}. "
            "Ожидается шаблон: <состав>_freq=<частота>_U=<напряжение>_<прочие параметры>.rwd"
        )
    frequency = match.group("frequency")
    is_batch = "from" in frequency.lower() and "to" in frequency.lower()
    if is_batch:
        return ExperimentIdentity(
            block_name=f"{match.group('prefix')}_{match.group('voltage')}",
            voltage=match.group("voltage"),
            frequency=frequency.strip("_"),
            frequency_sort_key=None,
            is_batch=True,
        )

    frequency_number = frequency.lower().removesuffix("hz").replace(",", ".")
    try:
        if "." not in frequency_number and len(frequency_number) > 1 and frequency_number.startswith("0"):
            sort_key = float(f"0.{frequency_number[1:]}")
        else:
            sort_key = float(frequency_number)
    except ValueError as error:
        raise RuntimeError(f"Не удалось определить числовую частоту из имени файла {path.name!r}.") from error
    return ExperimentIdentity(
        block_name=f"{match.group('prefix')}_{match.group('voltage')}",
        voltage=match.group("voltage"),
        frequency=frequency,
        frequency_sort_key=sort_key,
        is_batch=False,
    )


def group_records(records: list[RwdRecord]) -> list[tuple[str, list[tuple[ExperimentIdentity, RwdRecord]]]]:
    grouped: dict[str, list[tuple[ExperimentIdentity, RwdRecord]]] = {}
    seen: set[tuple[str, str]] = set()
    for record in records:
        identity = parse_experiment_identity(record.path)
        duplicate_key = (
            identity.block_name.lower(),
            "__batch__" if identity.is_batch else identity.frequency.lower(),
        )
        if duplicate_key in seen:
            raise RuntimeError(
                f"Для блока {identity.block_name!r} передано несколько файлов частоты {identity.frequency!r}. "
                "Оставьте один файл или уточните правило выбора версии."
            )
        seen.add(duplicate_key)
        grouped.setdefault(identity.block_name, []).append((identity, record))
    return [
        (
            block_name,
            sorted(
                items,
                key=lambda item: (
                    item[0].frequency_sort_key if item[0].frequency_sort_key is not None else -1.0,
                    item[1].path.name.lower(),
                ),
            ),
        )
        for block_name, items in sorted(grouped.items(), key=lambda item: item[0].lower())
    ]


def next_append_row(ws) -> int:
    last_nonempty = 0
    for row in range(ws.max_row, 0, -1):
        if any(ws.cell(row, column).value is not None for column in range(1, ws.max_column + 1)):
            last_nonempty = row
            break
    return 1 if last_nonempty == 0 else last_nonempty + 3


def existing_block_names(ws) -> set[str]:
    return {
        str(ws.cell(row, 2).value).strip().lower()
        for row in range(1, ws.max_row + 1)
        if ws.cell(row, 2).value is not None
    }


def template_voltage_rows(ws) -> dict[str, int]:
    rows: dict[str, int] = {}
    for row in range(1, ws.max_row + 1):
        value = ws.cell(row, 2).value
        if not isinstance(value, str):
            continue
        match = re.search(r"_U=(\d+v)_", value, re.I)
        if match:
            rows[match.group(1).lower()] = row
    return rows


def template_frequency_columns(ws, block_row: int) -> dict[float, int]:
    columns: dict[float, int] = {}
    for column in range(1, ws.max_column + 1):
        value = ws.cell(block_row + 1, column).value
        if not isinstance(value, str):
            continue
        match = re.search(r"Частота\s+([0-9.,]+)\s*Гц", value, re.I)
        if match:
            columns[float(match.group(1).replace(",", "."))] = column
    return columns


def find_frequency_column(columns: dict[float, int], frequency: float) -> int:
    for available_frequency, column in columns.items():
        if math.isclose(available_frequency, frequency, rel_tol=1e-9, abs_tol=1e-9):
            return column
    nearest_frequency = min(
        columns,
        key=lambda available_frequency: abs(available_frequency - frequency),
        default=None,
    )
    if nearest_frequency is not None and math.isclose(
        nearest_frequency,
        frequency,
        rel_tol=0.05,
        abs_tol=0.05,
    ):
        return columns[nearest_frequency]
    raise RuntimeError(
        f"В Excel-шаблоне отсутствует подблок частоты {frequency:g} Гц."
    )


def filename_frequency_note(identity: ExperimentIdentity, frequencies: list[float]) -> str:
    if identity.is_batch or identity.frequency_sort_key is None or not frequencies:
        return ""
    if any(
        math.isclose(identity.frequency_sort_key, frequency, rel_tol=1e-9, abs_tol=1e-9)
        for frequency in frequencies
    ):
        return ""
    found = ", ".join(f"{frequency:g} Гц" for frequency in sorted(frequencies))
    return (
        f" ВНИМАНИЕ: имя файла содержит freq={identity.frequency} "
        f"({identity.frequency_sort_key:g} Гц), но канал f in Hz содержит: {found}."
    )


def partial_segment_note(segment_counts: list[tuple[float, int, int]]) -> str:
    partial = [
        f"{frequency:g} Гц: {matched} из {expected}"
        for frequency, matched, expected in segment_counts
        if matched < expected
    ]
    if not partial:
        return ""
    return " ВНИМАНИЕ: записаны не все точки сегмента: " + "; ".join(partial) + "."


def validate_record_frequency(
    record: RwdRecord,
    frequency: float,
    row_count: int,
    bindings: list[ChannelBinding],
    start_index: int = 0,
    segment_index: int | None = None,
) -> None:
    frequency_binding = next(
        (
            binding
            for binding in bindings
            if normalize_header(binding.header) == normalize_header("f in Hz")
        ),
        None,
    )
    if frequency_binding is None:
        raise RuntimeError("В ASCII структуры отсутствует обязательная колонка 'f in Hz'.")
    actual_values = resolve_values_slice(
        record,
        frequency_binding,
        start_index,
        row_count,
        segment_index,
    )
    if not all(value_matches(frequency, item.value) for item in actual_values):
        raise RuntimeError(
            f"Файл {record.path.name}: канал частоты не соответствует имени файла "
            f"или ASCII-сегменту ({frequency:g} Гц). Для этой структуры нужен отдельный ASCII-файл."
        )


def fill_template_measurement_blocks(
    ws,
    records: list[RwdRecord],
    ascii_record: AsciiRecord,
    bindings: list[ChannelBinding],
) -> list[str] | None:
    voltage_rows = template_voltage_rows(ws)
    if not voltage_rows:
        return None

    segments = ascii_segments(ascii_record)
    seen: set[tuple[str, float]] = set()
    touched_voltages: set[str] = set()
    written_by_voltage: dict[str, set[float]] = {}
    report: list[str] = []
    for record in records:
        identity = parse_experiment_identity(record.path)
        record_bindings = adapt_bindings_to_record(bindings, record)
        selected_bindings = display_bindings(record_bindings)
        voltage = identity.voltage.removeprefix("U=").lower()
        touched_voltages.add(voltage)
        block_row = voltage_rows.get(voltage)
        if block_row is None:
            raise RuntimeError(
                f"В Excel-шаблоне отсутствует блок напряжения {identity.voltage!r}."
            )
        if identity.is_batch or len(segments) > 1:
            record_segments = [
                (ascii_segment_index, record_segment_index, segments[ascii_segment_index], matched_row_count)
                for ascii_segment_index, record_segment_index, matched_row_count in available_ascii_segment_indexes(
                    record,
                    record_bindings,
                    segments,
                )
            ]
        else:
            record_segments = [
                (
                    0,
                    0,
                    AsciiSegment(
                        label=segments[0].label,
                        start_index=segments[0].start_index,
                        row_count=segments[0].row_count,
                        frequency_sort_key=identity.frequency_sort_key,
                        frequency_label=identity.frequency,
                    ),
                    segments[0].row_count,
                )
            ]

        written_frequencies: list[float] = []
        segment_counts: list[tuple[float, int, int]] = []
        for ascii_segment_index, record_segment_index, segment, matched_row_count in record_segments:
            if segment.frequency_sort_key is None:
                raise RuntimeError(f"Не удалось определить частоту для файла {record.path.name!r}.")
            duplicate_key = (voltage, segment.frequency_sort_key)
            if duplicate_key in seen:
                raise RuntimeError(
                    f"Для напряжения {identity.voltage!r} передано несколько файлов "
                    f"или сегментов частоты {segment.frequency_sort_key:g} Гц."
                )
            seen.add(duplicate_key)
            validate_record_frequency(
                record,
                segment.frequency_sort_key,
                matched_row_count,
                record_bindings,
                segment.start_index,
                record_segment_index,
            )

            frequency_start_column = find_frequency_column(
                template_frequency_columns(ws, block_row),
                segment.frequency_sort_key,
            )
            rows = extract_signed_rows(
                record,
                matched_row_count,
                selected_bindings,
                segment.start_index,
                record_segment_index,
            )
            for row_offset, values in enumerate(rows, start=3):
                ws.cell(block_row + row_offset, frequency_start_column - 1, values[0])
                for column_offset, value in enumerate(values[1:]):
                    ws.cell(block_row + row_offset, frequency_start_column + column_offset, value)
            written_frequencies.append(segment.frequency_sort_key)
            written_by_voltage.setdefault(voltage, set()).add(segment.frequency_sort_key)
            segment_counts.append((segment.frequency_sort_key, matched_row_count, segment.row_count))
        frequencies = ", ".join(f"{frequency:g} Гц" for frequency in sorted(written_frequencies))
        report.append(
            f"{record.path.name}: блок {identity.voltage}, частоты: {frequencies}"
            + filename_frequency_note(identity, written_frequencies)
            + partial_segment_note(segment_counts)
        )
    for voltage in sorted(touched_voltages):
        block_row = voltage_rows[voltage]
        template_frequencies = sorted(template_frequency_columns(ws, block_row))
        written = written_by_voltage.get(voltage, set())
        missing = [
            frequency
            for frequency in template_frequencies
            if not any(math.isclose(frequency, item, rel_tol=0.05, abs_tol=0.05) for item in written)
        ]
        if missing:
            report.append(
                f"ВНИМАНИЕ: блок U={voltage}: не заполнены частоты шаблона: "
                + ", ".join(f"{frequency:g} Гц" for frequency in missing)
                + ". В переданных .rwd нет таких фактических частот."
            )
    return report


def append_measurement_blocks(
    ws,
    records: list[RwdRecord],
    ascii_record: AsciiRecord,
    bindings: list[ChannelBinding],
) -> list[str]:
    template_report = fill_template_measurement_blocks(ws, records, ascii_record, bindings)
    if template_report is not None:
        return template_report

    segments = ascii_segments(ascii_record)
    groups = group_records(records)
    existing_names = existing_block_names(ws)
    duplicates = [block_name for block_name, _ in groups if block_name.lower() in existing_names]
    if duplicates:
        raise RuntimeError(
            "В выбранном листе уже существуют блоки: "
            + ", ".join(duplicates)
            + ". Удалите старые блоки или используйте другой шаблон."
    )
    fill = PatternFill("solid", fgColor="1F4E78")
    block_start_row = next_append_row(ws)
    report: list[str] = []

    for block_name, items in groups:
        ws.cell(block_start_row, 2, block_name)
        subblocks: list[
            tuple[
                float,
                str,
                RwdRecord,
                list[ChannelBinding],
                list[tuple[str, ChannelBinding]],
                int,
                int,
                int | None,
            ]
        ] = []
        for identity, record in items:
            record_bindings = adapt_bindings_to_record(bindings, record)
            selected_bindings = display_bindings(record_bindings)
            if identity.is_batch or len(segments) > 1:
                subblocks.extend(
                    (
                        segment.frequency_sort_key,
                        segment.frequency_label,
                        record,
                        record_bindings,
                        selected_bindings,
                        segment.start_index,
                        row_count,
                        record_segment_index,
                    )
                    for ascii_segment_index, record_segment_index, matched_row_count in available_ascii_segment_indexes(
                        record,
                        record_bindings,
                        segments,
                    )
                    for segment in [segments[ascii_segment_index]]
                    for row_count in [matched_row_count]
                )
            else:
                if identity.frequency_sort_key is None:
                    raise RuntimeError(f"Не удалось определить частоту из имени файла {record.path.name!r}.")
                subblocks.append(
                    (
                        identity.frequency_sort_key,
                        identity.frequency,
                        record,
                        record_bindings,
                        selected_bindings,
                        segments[0].start_index,
                        segments[0].row_count,
                        0 if bindings and bindings[0].segment_offsets else None,
                    )
                )

        duplicate_frequencies = {
            frequency
            for frequency in [item[0] for item in subblocks]
            if sum(1 for other in subblocks if math.isclose(other[0], frequency, rel_tol=1e-9, abs_tol=1e-9)) > 1
        }
        if duplicate_frequencies:
            raise RuntimeError(
                f"Для блока {block_name!r} передано несколько данных одной частоты: "
                + ", ".join(f"{frequency:g} Гц" for frequency in sorted(duplicate_frequencies))
            )

        max_row_count = max(row_count for _, _, _, _, _, _, row_count, _ in subblocks)
        for frequency_index, (
            frequency,
            frequency_label,
            record,
            record_bindings,
            selected_bindings,
            start_index,
            row_count,
            segment_index,
        ) in enumerate(
            sorted(subblocks, key=lambda item: item[0])
        ):
            frequency_start_column = 1 + frequency_index * 7
            value_start_column = frequency_start_column + 1
            validate_record_frequency(record, frequency, row_count, record_bindings, start_index, segment_index)
            ws.cell(block_start_row + 1, value_start_column, frequency_label)
            for column_offset, (display_header, _) in enumerate(selected_bindings):
                ws.cell(block_start_row + 2, value_start_column + column_offset, display_header)

            rows = extract_signed_rows(record, row_count, selected_bindings, start_index, segment_index)
            for row_offset, row in enumerate(rows, start=3):
                ws.cell(block_start_row + row_offset, frequency_start_column, row[0])
                for column_offset, value in enumerate(row[1:]):
                    ws.cell(block_start_row + row_offset, value_start_column + column_offset, value)

            for row in (block_start_row + 1, block_start_row + 2):
                for column in range(frequency_start_column, frequency_start_column + 7):
                    cell = ws.cell(row, column)
                    if cell.value is not None:
                        cell.font = Font(bold=True, color="FFFFFF")
                        cell.fill = fill
                        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            for column in range(frequency_start_column, frequency_start_column + 7):
                ws.column_dimensions[get_column_letter(column)].width = 18
        report_frequencies = ", ".join(
            f"{frequency:g} Гц" for frequency, *_ in sorted(subblocks, key=lambda item: item[0])
        )
        notes = [
            filename_frequency_note(identity, [frequency])
            for frequency, _, record, _, _, _, _, _ in subblocks
            for identity in [parse_experiment_identity(record.path)]
        ]
        report.append(
            f"Добавлен новый блок {block_name}: частоты: {report_frequencies}"
            + "".join(note for note in notes if note)
        )
        block_start_row += max_row_count + 5
    ws.freeze_panes = "B4"
    return report


def metadata_fields(record: RwdRecord) -> dict[str, Any]:
    return {
        "file_path": str(record.path),
        "size_bytes": record.size_bytes,
        "format": record.format_magic,
        "rheowin_version": record.rheowin_version,
        "date": record.date,
        "time": record.time,
        "instrument_model": record.instrument_model,
        "device": record.device,
        "geometry": record.geometry,
        "temperature_controller": record.temperature_controller,
        "gap": record.gap,
        "serial_numbers": record.serial_numbers,
        "driver_versions": record.driver_versions,
        "firmware_versions": record.firmware_versions,
        "job_file": record.job_file,
        "mode": record.mode,
        "numeric_blocks_extracted": len(record.numeric_blocks),
    }


def update_metadata_sheet(workbook, records: list[RwdRecord]) -> None:
    sheet_name = "instrument_metadata"
    ws = workbook[sheet_name] if sheet_name in workbook.sheetnames else workbook.create_sheet(sheet_name)
    existing: dict[str, dict[str, Any]] = {}
    if ws.max_row >= 1 and ws.max_column >= 2:
        fields = [ws.cell(row, 1).value for row in range(2, ws.max_row + 1)]
        for column in range(2, ws.max_column + 1):
            file_name = ws.cell(1, column).value
            if file_name:
                existing[str(file_name)] = {
                    str(field): ws.cell(row, column).value
                    for row, field in enumerate(fields, start=2)
                    if field is not None
                }
    for record in records:
        existing[record.path.name] = metadata_fields(record)

    field_names = list(metadata_fields(records[0]).keys())
    ws.delete_rows(1, ws.max_row)
    ws.delete_cols(1, ws.max_column)
    ws.cell(1, 1, "field")
    for column, file_name in enumerate(sorted(existing, key=str.lower), start=2):
        ws.cell(1, column, file_name)
        for row, field_name in enumerate(field_names, start=2):
            ws.cell(row, 1, field_name)
            ws.cell(row, column, existing[file_name].get(field_name, ""))
    style_table(ws)


def style_table(ws) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for column in range(1, ws.max_column + 1):
        width = max(len(str(ws.cell(row, column).value or "")) for row in range(1, min(ws.max_row, 500) + 1))
        ws.column_dimensions[get_column_letter(column)].width = min(max(width + 2, 12), 70)


def write_mapping_sheet(workbook, ascii_record: AsciiRecord, bindings: list[ChannelBinding]) -> None:
    sheet_name = "parser_mapping"
    if sheet_name in workbook.sheetnames:
        del workbook[sheet_name]
    ws = workbook.create_sheet(sheet_name)
    rows = [
        ["structure_ascii", str(ascii_record.path)],
        ["mapping_mode", "structural"],
        [],
        ["header", "code_1", "code_2", "dtype", "occurrence", "segment_stride"],
    ]
    rows.extend(
        [
            [
                binding.header,
                binding.block_key[0] if binding.block_key else "",
                binding.block_key[1] if binding.block_key else "",
                binding.block_key[2] if binding.block_key else "",
                binding.block_occurrence,
                binding.segment_stride,
            ]
            for binding in bindings
        ]
    )
    for row in rows:
        ws.append(row)
    style_table(ws)
    ws.sheet_state = "hidden"


def write_diagnostics_sheet(workbook, records: list[RwdRecord]) -> None:
    sheet_name = "raw_values"
    if sheet_name in workbook.sheetnames:
        del workbook[sheet_name]
    ws = workbook.create_sheet(sheet_name)
    ws.append(
        [
            "file_name",
            "record_offset_hex",
            "code_1",
            "code_2",
            "dtype",
            "value_index",
            "value_offset_hex",
            "raw_hex_le",
            "float32_value",
        ]
    )
    for record in records:
        for block in record.numeric_blocks:
            for value_index, item in enumerate(block.values, start=1):
                ws.append(
                    [
                        record.path.name,
                        hex(block.offset),
                        block.code_1,
                        block.code_2,
                        block.dtype,
                        value_index,
                        hex(item.offset),
                        item.raw_hex_le,
                        excel_value(item.value),
                    ]
                )
    style_table(ws)


def output_path_for(template: Path, requested: Path | None) -> Path:
    if requested:
        return requested.expanduser().resolve()
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return template.with_name(f"{template.stem}_parsed_{timestamp}.xlsx")


def parse_dragged_paths(raw_value: str) -> list[Path]:
    cleaned = raw_value.strip()
    if not cleaned:
        return []
    try:
        parts = shlex.split(cleaned)
    except ValueError:
        parts = [cleaned.strip("'\"")]
    return [Path(part).expanduser() for part in parts]


def prompt_line(prompt: str, default: str | None = None) -> str:
    suffix = f" [{default}]" if default else ""
    value = input(f"{prompt}{suffix}: ").strip()
    return value or (default or "")


def prompt_existing_file(prompt: str, suffixes: set[str]) -> Path:
    while True:
        paths = parse_dragged_paths(prompt_line(prompt))
        if len(paths) != 1:
            print("Укажите ровно один файл. Можно перетащить файл в терминал.")
            continue
        path = paths[0].resolve()
        if not path.is_file():
            print(f"Файл не найден: {path}")
            continue
        if path.suffix.lower() not in suffixes:
            print(f"Нужен файл с расширением: {', '.join(sorted(suffixes))}")
            continue
        return path


def prompt_rwd_inputs() -> list[Path]:
    while True:
        paths = parse_dragged_paths(
            prompt_line(
                "3. Укажите один или несколько .rwd файлов или папку с .rwd"
            )
        )
        if not paths:
            print("Нужно указать хотя бы один .rwd файл или папку.")
            continue
        return [path.resolve() for path in paths]


def prompt_yes_no(prompt: str, default: bool = False) -> bool:
    default_text = "y" if default else "n"
    value = prompt_line(prompt, default_text).strip().lower()
    return value in {"y", "yes", "д", "да"}


def prompt_output_path(template: Path) -> Path:
    default = output_path_for(template, None)
    while True:
        raw_value = prompt_line("5. Укажите итоговый .xlsx файл", str(default))
        paths = parse_dragged_paths(raw_value)
        path = (paths[0] if paths else Path(raw_value)).expanduser().resolve()
        if path.suffix.lower() != ".xlsx":
            print("Итоговый файл должен иметь расширение .xlsx.")
            continue
        if path == template:
            print(
                "Вы выбрали тот же файл, что и шаблон. Лучше писать в новый .xlsx, "
                "особенно если шаблон открыт в Excel."
            )
            if not prompt_yes_no("Все равно перезаписать этот файл?", default=False):
                continue
        return path


def interactive_args() -> argparse.Namespace:
    print("HAAKE .rwd -> Excel")
    print("Можно перетаскивать файлы и папки в терминал мышкой.")
    template = prompt_existing_file("1. Укажите Excel-шаблон .xlsx", {".xlsx"})
    ascii_path = prompt_existing_file(
        "2. Укажите ASCII-файл структуры RheoWin .txt/.asc/.csv",
        ASCII_SUFFIXES,
    )
    rwd_paths = prompt_rwd_inputs()
    recursive = prompt_yes_no("Искать .rwd во вложенных папках?", default=False)
    sheet = prompt_line("4. Укажите лист, куда писать данные", "input_data")
    output = prompt_output_path(template)
    include_diagnostics = prompt_yes_no("Добавить диагностический лист raw_values?", default=False)
    return argparse.Namespace(
        template=template,
        ascii=ascii_path,
        rwd=rwd_paths,
        sheet=sheet,
        output=output,
        recursive=recursive,
        include_diagnostics=include_diagnostics,
    )


def parse_args(argv: list[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Заполнить Excel-шаблон фактическими данными из HAAKE RheoWin .rwd по структуре ASCII."
    )
    parser.add_argument("template", type=Path, help="Первый файл: Excel-шаблон .xlsx.")
    parser.add_argument("ascii", type=Path, help="Второй файл: ASCII-экспорт RheoWin со структурой колонок .txt/.asc/.csv.")
    parser.add_argument("rwd", nargs="+", type=Path, help="Один или несколько .rwd файлов либо папок с .rwd.")
    parser.add_argument("--sheet", required=True, help="Название листа Excel, в который добавляются группы данных.")
    parser.add_argument("-o", "--output", type=Path, help="Путь к новой итоговой книге .xlsx.")
    parser.add_argument("-r", "--recursive", action="store_true", help="Искать .rwd во вложенных папках.")
    parser.add_argument("--include-diagnostics", action="store_true", help="Добавить технический лист raw_values.")
    return parser.parse_args(argv)


def validate_inputs(args: argparse.Namespace) -> tuple[Path, Path, list[Path]]:
    template = args.template.expanduser().resolve()
    ascii_path = args.ascii.expanduser().resolve()
    if not template.is_file() or template.suffix.lower() != ".xlsx":
        raise RuntimeError("Первым аргументом должен быть существующий Excel-шаблон .xlsx.")
    if not ascii_path.is_file() or ascii_path.suffix.lower() not in ASCII_SUFFIXES:
        raise RuntimeError("Вторым аргументом должен быть существующий ASCII-файл .txt, .asc или .csv.")
    rwd_files = collect_rwd_files(args.rwd, recursive=args.recursive)
    if not rwd_files:
        raise RuntimeError("Не найдены .rwd файлы для обработки.")
    return template, ascii_path, rwd_files


def main(argv: list[str] | None = None) -> int:
    timer = StepTimer()
    try:
        argv = sys.argv[1:] if argv is None else argv
        args = interactive_args() if not argv else parse_args(argv)
        template, ascii_path, rwd_files = validate_inputs(args)
        timer.mark("Валидация входных путей")
        ascii_record = extract_ascii_record(ascii_path)
        timer.mark("Чтение ASCII-структуры")
        records = [extract_rwd_record(path) for path in rwd_files]
        timer.mark("Чтение .rwd файлов")
        bindings = build_structural_bindings(ascii_record, records)
        timer.mark("Построение структурной карты")

        workbook = load_workbook(template)
        if args.sheet not in workbook.sheetnames:
            raise RuntimeError(
                f"В шаблоне нет листа {args.sheet!r}. Доступные листы: {', '.join(workbook.sheetnames)}"
            )
        timer.mark("Загрузка Excel-шаблона")
        fill_report = append_measurement_blocks(workbook[args.sheet], records, ascii_record, bindings)
        update_metadata_sheet(workbook, records)
        write_mapping_sheet(workbook, ascii_record, bindings)
        if args.include_diagnostics:
            write_diagnostics_sheet(workbook, records)
        timer.mark("Заполнение Excel-книги")

        output = output_path_for(template, args.output)
        workbook.calculation.fullCalcOnLoad = True
        workbook.calculation.forceFullCalc = True
        workbook.calculation.calcMode = "auto"
        workbook.save(output)
        timer.mark("Сохранение Excel-книги")
        print(f"ASCII-структура: {ascii_record.path.name}")
        print(f"Структурно сопоставлено колонок: {len(bindings)}")
        print(f"Обработано .rwd файлов: {len(records)}")
        print("Записано в Excel:")
        for line in fill_report:
            print(f"  - {line}")
        print(f"Создан файл: {output}")
        print("Время выполнения:")
        for label, seconds in timer.steps:
            print(f"  - {label}: {format_elapsed(seconds)}")
        print(f"  - Итого: {format_elapsed(timer.total)}")
        return 0
    except RuntimeError as error:
        print(f"Ошибка: {error}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
