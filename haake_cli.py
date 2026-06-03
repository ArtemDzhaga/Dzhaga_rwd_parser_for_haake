#!/usr/bin/env python3
"""
Локальный CLI для переноса данных HAAKE RheoWin .rwd в Excel-шаблон.

Порядок аргументов:
1. Excel-шаблон .xlsx.
2. Эталонный ASCII-экспорт RheoWin .txt/.asc/.csv.
3. Один или несколько .rwd файлов либо папок с .rwd.

ASCII используется как доказательная калибровка: скрипт находит среди
переданных .rwd соответствующий эталонному ASCII файл и сопоставляет каждый
официально подписанный столбец конкретному бинарному каналу значение-в-значение.
Если доказательное сопоставление невозможно, скрипт останавливается.
"""

from __future__ import annotations

import argparse
import csv
import math
import re
import struct
import sys
from copy import deepcopy
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path, PureWindowsPath
from typing import Any

from openpyxl import load_workbook
from openpyxl.chart.data_source import AxDataSource, NumDataSource, NumRef
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.chart.legend import Legend
from openpyxl.chart.series import SeriesLabel
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.line import LineProperties
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


NUMERIC_RECORD_MARKER = bytes.fromhex("da00da00feff")
TEXT_RE = re.compile(rb"[\x09\x0a\x0d\x20-\x7e]{4,}")
ASCII_SUFFIXES = {".txt", ".asc", ".csv"}
CHART_VOLTAGE_STYLES = [
    ("0v", "404040", "square"),
    ("500v", "E53935", "diamond"),
    ("1000v", "2878D4", "triangle"),
    ("1500v", "43A047", "circle"),
    ("2000v", "7E57C2", "x"),
    ("2500v", "FB8C00", "star"),
]
CHART_METRIC_COLUMNS = [
    ([3, 5, 7, 9, 11, 13], [4, 6, 8, 10, 12, 14]),
    ([16, 18, 20, 22, 24, 26], [17, 19, 21, 23, 25, 27]),
    ([29, 31, 33, 35, 37, 39], [30, 32, 34, 36, 38, 40]),
    ([3, 5, 7, 9, 11, 13], [68, 69, 70, 71, 72, 73]),
]
CHART_METRIC_INPUT_OFFSETS = [2, 3, 4, None]
CHART_CURVE_POINT_COUNT = 81
CHART_POLYNOMIAL_ORDER = 3


@dataclass
class NumericValue:
    value: float
    raw_hex_le: str
    offset: int


@dataclass
class NumericBlock:
    offset: int
    data_offset: int
    code_1: int
    code_2: int
    dtype: int
    values: list[NumericValue]
    extended_values: list[NumericValue]

    @property
    def key(self) -> tuple[int, int, int]:
        return self.code_1, self.code_2, self.dtype


@dataclass
class RwdRecord:
    path: Path
    blob: bytes
    size_bytes: int
    format_magic: str
    rheowin_version: str
    date: str
    time: str
    device: str
    instrument_model: str
    geometry: str
    temperature_controller: str
    gap: str
    job_file: str
    mode: str
    serial_numbers: str
    driver_versions: str
    firmware_versions: str
    strings: list[tuple[int, str]]
    numeric_blocks: list[NumericBlock]


@dataclass
class AsciiRecord:
    path: Path
    source_rwd_path: str
    operator: str
    date_time_version: str
    headers: list[str]
    rows: list[list[str | float]]

    @property
    def row_labels(self) -> list[str]:
        return [str(row[0]) for row in self.rows]


@dataclass
class ChannelBinding:
    header: str
    raw_offset: int


@dataclass(frozen=True)
class ExperimentIdentity:
    block_name: str
    voltage: str
    frequency: str
    frequency_sort_key: float


DISPLAY_COLUMNS = [
    ("t_seg in s", "t_seg in s"),
    ("Tau in Pa", "Tau in Pa"),
    ("G' in Pa", "G' in Pa"),
    ('G" in Pa', 'G" in Pa'),
    ("|Eta*| in Pas", "|eta*| in Pas"),
    ("Gamma in -", "gamma"),
]


def extract_ascii_strings(blob: bytes) -> list[tuple[int, str]]:
    rows: list[tuple[int, str]] = []
    for match in TEXT_RE.finditer(blob):
        text = match.group().decode("latin1", errors="replace").strip()
        if text:
            rows.append((match.start(), text))
    return rows


def extract_numeric_blocks(blob: bytes) -> list[NumericBlock]:
    blocks: list[NumericBlock] = []
    start = 0
    while True:
        offset = blob.find(NUMERIC_RECORD_MARKER, start)
        if offset < 0:
            break
        start = offset + 1

        if offset + 60 > len(blob):
            continue

        code_1 = struct.unpack_from("<H", blob, offset + 6)[0]
        code_2 = struct.unpack_from("<H", blob, offset + 8)[0]
        dtype = struct.unpack_from("<I", blob, offset + 14)[0]
        count = struct.unpack_from("<I", blob, offset + 18)[0]
        if dtype != 4 or count <= 0 or count > 10000:
            continue

        data_offset = offset + 60
        data_end = data_offset + count * 4
        if data_end > len(blob):
            continue

        # Some RheoWin job-manager files declare half of the physical array
        # length in the record header. Keep one extra declared-length chunk:
        # calibration accepts it only when it matches the ASCII export
        # value-for-value.
        extended_count = min(count * 2, (len(blob) - data_offset) // 4)
        extended_values: list[NumericValue] = []
        for index in range(extended_count):
            value_offset = data_offset + index * 4
            raw_bytes = blob[value_offset : value_offset + 4]
            extended_values.append(
                NumericValue(
                    value=struct.unpack("<f", raw_bytes)[0],
                    raw_hex_le=raw_bytes.hex(),
                    offset=value_offset,
                )
            )
        blocks.append(
            NumericBlock(
                offset=offset,
                data_offset=data_offset,
                code_1=code_1,
                code_2=code_2,
                dtype=dtype,
                values=extended_values[:count],
                extended_values=extended_values,
            )
        )
    return blocks


def unique_join(values: list[str]) -> str:
    seen: set[str] = set()
    result: list[str] = []
    for value in values:
        cleaned = value.strip()
        if cleaned and cleaned not in seen:
            seen.add(cleaned)
            result.append(cleaned)
    return " | ".join(result)


def first_match(pattern: str, text: str, flags: int = 0) -> str:
    match = re.search(pattern, text, flags)
    return match.group(1).strip() if match else ""


def all_matches(pattern: str, text: str, flags: int = 0) -> list[str]:
    return [match.strip() for match in re.findall(pattern, text, flags) if match.strip()]


def parse_pipe_metadata(text: str, key: str) -> list[str]:
    return all_matches(rf"{re.escape(key)}:([^|$\n\r]+)", text)


def detect_device(lines: list[str], text: str) -> str:
    candidates: list[str] = []
    for line in lines:
        if "\\" in line or ".pdf" in line.lower() or ".rwj" in line.lower():
            continue
        if re.search(r"\bRheoStress\b|\bRheoWin\b|\bRS1500?\b", line, re.I):
            if not line.startswith("***") and len(line) < 120:
                candidates.append(line)
    candidates.extend(all_matches(r"(RheoStress\s+RS\d+)", text, re.I))
    candidates.extend(all_matches(r"\b(RS\d{2,4})\b", text))
    return unique_join(candidates[:8])


def detect_geometry(lines: list[str], text: str) -> str:
    candidates: list[str] = []
    for line in lines:
        if re.search(r"\b(PP|CP|DG|Z|C)\d+\b|plate|cone|sensor", line, re.I) and len(line) < 120:
            candidates.append(line)
    candidates.extend(all_matches(r"\b(PP\d+\s*[A-Z0-9 ]*)", text))
    return unique_join(candidates[:8])


def extract_rwd_record(path: Path) -> RwdRecord:
    blob = path.read_bytes()
    strings = extract_ascii_strings(blob)
    lines = [text for _, text in strings]
    full_text = "\n".join(lines)
    date_time = re.search(r"Date/Time:\s*([0-9.]+)\s*/\s*([0-9:]+)", full_text)

    firmware_versions = parse_pipe_metadata(full_text, "Firmware version 1")
    firmware_versions.extend(parse_pipe_metadata(full_text, "Firmware version 3"))
    return RwdRecord(
        path=path,
        blob=blob,
        size_bytes=len(blob),
        format_magic="OOP_HADES" if b"OOP_HADES" in blob else blob[:16].hex(" "),
        rheowin_version=first_match(r"HAAKE RheoWin\s+([0-9.]+)", full_text),
        date=date_time.group(1) if date_time else first_match(r"\b([0-3]\d\.[01]\d\.\d{4})\b", full_text),
        time=date_time.group(2) if date_time else first_match(r"\b([0-2]\d:[0-5]\d:[0-5]\d)\b", full_text),
        device=detect_device(lines, full_text),
        instrument_model=unique_join(all_matches(r"\b(RheoStress\s+RS\d+|RS1500?|DC50)\b", full_text, re.I)),
        geometry=detect_geometry(lines, full_text),
        temperature_controller=unique_join(all_matches(r"\b(DC50\b[^\n\r]*)", full_text)),
        gap=first_match(r"Gap:\s*([^\n\r]+)", full_text),
        job_file=unique_join(all_matches(r"([A-Z]:\\[^\n\r]+?\.rwj)", full_text, re.I)),
        mode=first_match(r"ElmStros::Execute:\s*Mode:\s*([^\n\r]+)", full_text),
        serial_numbers=unique_join(parse_pipe_metadata(full_text, "Serial number")),
        driver_versions=unique_join(parse_pipe_metadata(full_text, "Driver version")),
        firmware_versions=unique_join(firmware_versions),
        strings=strings,
        numeric_blocks=extract_numeric_blocks(blob),
    )


def read_ascii_text(path: Path) -> str:
    for encoding in ("utf-8-sig", "cp1251", "cp1252", "latin1"):
        try:
            return path.read_text(encoding=encoding)
        except UnicodeError:
            continue
    raise RuntimeError(f"Не удалось прочитать ASCII-файл: {path}")


def parse_ascii_number(value: str) -> str | float:
    stripped = value.strip()
    if not stripped:
        return ""
    try:
        return float(stripped.replace(",", "."))
    except ValueError:
        return stripped


def extract_ascii_record(path: Path) -> AsciiRecord:
    lines = read_ascii_text(path).splitlines()
    header_index = next(
        (index for index, line in enumerate(lines) if line.startswith(";") and line.count(";") >= 2),
        None,
    )
    if header_index is None:
        raise RuntimeError(f"В ASCII-файле не найдена строка заголовков RheoWin: {path}")

    headers = next(csv.reader([lines[header_index]], delimiter=";"))
    headers = ["segment_point" if index == 0 and not header else header.strip() for index, header in enumerate(headers)]
    while headers and not headers[-1]:
        headers.pop()

    rows: list[list[str | float]] = []
    for line in lines[header_index + 1 :]:
        if not line.strip():
            continue
        cells = next(csv.reader([line], delimiter=";"))
        while cells and not cells[-1]:
            cells.pop()
        if len(cells) == len(headers):
            rows.append([cells[0].strip()] + [parse_ascii_number(cell) for cell in cells[1:]])

    if not rows:
        raise RuntimeError(f"В ASCII-файле нет строк измерений: {path}")

    metadata = [line.strip() for line in lines[:header_index] if line.strip()]
    return AsciiRecord(
        path=path,
        source_rwd_path=metadata[0] if metadata else "",
        operator=metadata[1] if len(metadata) > 1 else "",
        date_time_version=metadata[2] if len(metadata) > 2 else "",
        headers=headers,
        rows=rows,
    )


def value_matches(exported: float, actual: float) -> bool:
    if not math.isfinite(actual):
        return False
    tolerance = max(1e-7, abs(exported) * 0.001)
    return abs(exported - actual) <= tolerance


def column_match_score(exported: list[float], block: NumericBlock) -> float | None:
    if len(exported) == len(block.values):
        actual_values = block.values
    elif len(exported) == len(block.extended_values):
        actual_values = block.extended_values
    else:
        return None
    actual = [item.value for item in actual_values]
    if not all(value_matches(expected, observed) for expected, observed in zip(exported, actual)):
        return None
    return sum(abs(expected - observed) / max(1e-7, abs(expected) * 0.001) for expected, observed in zip(exported, actual))


def numeric_ascii_columns(ascii_record: AsciiRecord) -> list[tuple[int, str, list[float]]]:
    columns: list[tuple[int, str, list[float]]] = []
    for column_index, header in enumerate(ascii_record.headers[1:], start=1):
        values = [row[column_index] for row in ascii_record.rows]
        if all(isinstance(value, float) for value in values):
            columns.append((column_index, header, [float(value) for value in values]))
    return columns


def block_occurrence(blocks: list[NumericBlock], selected: NumericBlock) -> int:
    occurrence = 0
    for block in blocks:
        if block.key == selected.key:
            if block is selected:
                return occurrence
            occurrence += 1
    raise RuntimeError("Внутренняя ошибка: выбранный бинарный блок не найден.")


def raw_sequence_candidates(blob: bytes, exported_values: list[float]) -> list[int]:
    sequence_size = len(exported_values) * 4
    candidates: list[int] = []
    for offset in range(0, len(blob) - sequence_size + 1):
        first_value = struct.unpack_from("<f", blob, offset)[0]
        if not value_matches(exported_values[0], first_value):
            continue
        if all(
            value_matches(expected, struct.unpack_from("<f", blob, offset + index * 4)[0])
            for index, expected in enumerate(exported_values[1:], start=1)
        ):
            candidates.append(offset)
    return candidates


def required_ascii_columns(ascii_record: AsciiRecord) -> list[tuple[int, str, list[float]]]:
    required = {normalize_header(header) for header, _ in DISPLAY_COLUMNS}
    required.add(normalize_header("f in Hz"))
    return [
        column
        for column in numeric_ascii_columns(ascii_record)
        if normalize_header(column[1]) in required
    ]


def find_bindings(ascii_record: AsciiRecord, rwd_record: RwdRecord) -> tuple[list[ChannelBinding], list[str]]:
    bindings: list[ChannelBinding] = []
    errors: list[str] = []
    used_offsets: set[int] = set()
    for _, header, exported_values in required_ascii_columns(ascii_record):
        candidates = [
            offset
            for offset in raw_sequence_candidates(rwd_record.blob, exported_values)
            if offset not in used_offsets
        ]
        if not candidates:
            errors.append(f"{header}: совпадающий бинарный канал не найден")
            continue
        if len(candidates) > 1:
            errors.append(f"{header}: найдено несколько одинаково подходящих бинарных каналов")
            continue

        selected = candidates[0]
        used_offsets.add(selected)
        bindings.append(ChannelBinding(header=header, raw_offset=selected))
    return bindings, errors


def calibrate(ascii_record: AsciiRecord, records: list[RwdRecord]) -> tuple[RwdRecord, list[ChannelBinding]]:
    segment_numbers = {
        label.split("|", 1)[0]
        for label in ascii_record.row_labels
        if "|" in label
    }
    if len(segment_numbers) > 1:
        raise RuntimeError(
            f"ASCII содержит пакетный эксперимент: {len(ascii_record.rows)} строк "
            f"и {len(segment_numbers)} подблоков частот. Текущая версия CLI умеет "
            "доказательно калибровать только ASCII одного подблока. "
            "Для пакетных .rwd нужен отдельный режим разбора 10×30; не используйте "
            "этот ASCII как эталон одиночного файла."
        )

    attempts: list[tuple[RwdRecord, list[ChannelBinding], list[str]]] = []
    for record in records:
        bindings, errors = find_bindings(ascii_record, record)
        attempts.append((record, bindings, errors))

    expected_count = len(required_ascii_columns(ascii_record))
    complete = [attempt for attempt in attempts if len(attempt[1]) == expected_count and not attempt[2]]
    if len(complete) == 1:
        return complete[0][0], complete[0][1]
    if len(complete) > 1:
        names = ", ".join(record.path.name for record, _, _ in complete)
        raise RuntimeError(f"ASCII одинаково соответствует нескольким .rwd: {names}")

    best = max(attempts, key=lambda attempt: len(attempt[1]))
    details = "; ".join(best[2][:4]) or "совпадающие каналы не найдены"
    expected_name = PureWindowsPath(ascii_record.source_rwd_path).name
    expected_hint = (
        f" ASCII был экспортирован из {expected_name!r}; добавьте именно этот .rwd "
        "либо сделайте новый ASCII-экспорт из одного из переданных .rwd."
        if expected_name
        else ""
    )
    raise RuntimeError(
        "Не удалось доказательно сопоставить ASCII ни одному переданному .rwd. "
        f"Лучший кандидат: {best[0].path.name}; сопоставлено {len(best[1])} из {expected_count} колонок. "
        f"Причина: {details}.{expected_hint}"
    )


def resolve_values(record: RwdRecord, binding: ChannelBinding, row_count: int) -> list[NumericValue]:
    end_offset = binding.raw_offset + row_count * 4
    if end_offset > len(record.blob):
        raise RuntimeError(
            f"Файл {record.path.name}: отсутствует полный массив канала {binding.header} "
            f"по доказанному смещению {hex(binding.raw_offset)}."
        )
    values: list[NumericValue] = []
    for offset in range(binding.raw_offset, end_offset, 4):
        raw_bytes = record.blob[offset : offset + 4]
        values.append(
            NumericValue(
                value=struct.unpack("<f", raw_bytes)[0],
                raw_hex_le=raw_bytes.hex(),
                offset=offset,
            )
        )
    return values


def normalize_header(header: str) -> str:
    return re.sub(r"\s+", " ", header.strip()).lower().replace("''", '"')


def display_bindings(bindings: list[ChannelBinding]) -> list[tuple[str, ChannelBinding]]:
    by_header = {normalize_header(binding.header): binding for binding in bindings}
    result: list[tuple[str, ChannelBinding]] = []
    missing: list[str] = []
    for ascii_header, display_header in DISPLAY_COLUMNS:
        binding = by_header.get(normalize_header(ascii_header))
        if binding is None:
            missing.append(ascii_header)
        else:
            result.append((display_header, binding))
    if missing:
        raise RuntimeError(
            "В эталонном ASCII отсутствуют обязательные рабочие колонки: "
            + ", ".join(missing)
        )
    return result


def extract_signed_rows(
    record: RwdRecord,
    row_count: int,
    selected_bindings: list[tuple[str, ChannelBinding]],
) -> list[list[str | float]]:
    value_sets = [resolve_values(record, binding, row_count) for _, binding in selected_bindings]
    rows: list[list[str | float]] = []
    for index in range(row_count):
        rows.append(
            [f"1|{index + 1}"]
            + [excel_value(values[index].value) for values in value_sets]
        )
    return rows


def excel_value(value: float) -> float | str:
    if math.isnan(value):
        return "NaN"
    if math.isinf(value):
        return "+Inf" if value > 0 else "-Inf"
    return value


def collect_rwd_files(paths: list[Path], recursive: bool) -> list[Path]:
    files: list[Path] = []
    for path in paths:
        expanded = path.expanduser()
        if expanded.is_dir():
            files.extend(expanded.glob("**/*.rwd" if recursive else "*.rwd"))
        elif expanded.is_file() and expanded.suffix.lower() == ".rwd":
            files.append(expanded)
        else:
            visible_path = str(path).replace("\n", "\\n")
            raise RuntimeError(
                f"Не найден переданный .rwd файл или папка: {visible_path!r}. "
                "Проверьте, что путь не содержит перенос строки внутри кавычек."
            )
    return sorted(set(file.resolve() for file in files), key=lambda path: path.name.lower())


def parse_experiment_identity(path: Path) -> ExperimentIdentity:
    if "from_" in path.stem.lower() and "_to_" in path.stem.lower():
        raise RuntimeError(
            f"Файл {path.name!r} содержит пакет частот. "
            "Его нельзя смешивать с одиночными .rwd в текущем режиме 1×30. "
            "Для него нужен отдельный пакетный режим и соответствующий ASCII-экспорт 10×30."
        )
    match = re.match(
        r"^(?P<prefix>.+?)_freq=(?P<frequency>[^_]+)_(?P<voltage>U=[^_]+)(?:_.*)?$",
        path.stem,
        re.I,
    )
    if not match:
        raise RuntimeError(
            f"Не удалось разобрать имя файла {path.name!r}. "
            "Ожидается шаблон: <состав>_freq=<частота>_U=<напряжение>_<прочие параметры>.rwd"
        )
    frequency = match.group("frequency")
    frequency_number = frequency.lower().removesuffix("hz").replace(",", ".")
    try:
        if "." not in frequency_number and len(frequency_number) > 1 and frequency_number.startswith("0"):
            sort_key = float(f"0.{frequency_number[1:]}")
        else:
            sort_key = float(frequency_number)
    except ValueError as error:
        raise RuntimeError(f"Не удалось определить числовую частоту из имени файла {path.name!r}.") from error
    return ExperimentIdentity(
        block_name=f"{match.group('prefix')}_{match.group('voltage')}",
        voltage=match.group("voltage"),
        frequency=frequency,
        frequency_sort_key=sort_key,
    )


def group_records(records: list[RwdRecord]) -> list[tuple[str, list[tuple[ExperimentIdentity, RwdRecord]]]]:
    grouped: dict[str, list[tuple[ExperimentIdentity, RwdRecord]]] = {}
    seen: set[tuple[str, str]] = set()
    for record in records:
        identity = parse_experiment_identity(record.path)
        duplicate_key = (identity.block_name.lower(), identity.frequency.lower())
        if duplicate_key in seen:
            raise RuntimeError(
                f"Для блока {identity.block_name!r} передано несколько файлов частоты {identity.frequency!r}. "
                "Оставьте один файл или уточните правило выбора версии."
            )
        seen.add(duplicate_key)
        grouped.setdefault(identity.block_name, []).append((identity, record))
    return [
        (block_name, sorted(items, key=lambda item: (item[0].frequency_sort_key, item[1].path.name.lower())))
        for block_name, items in sorted(grouped.items(), key=lambda item: item[0].lower())
    ]


def next_append_row(ws) -> int:
    last_nonempty = 0
    for row in range(ws.max_row, 0, -1):
        if any(ws.cell(row, column).value is not None for column in range(1, ws.max_column + 1)):
            last_nonempty = row
            break
    return 1 if last_nonempty == 0 else last_nonempty + 3


def existing_block_names(ws) -> set[str]:
    return {
        str(ws.cell(row, 2).value).strip().lower()
        for row in range(1, ws.max_row + 1)
        if ws.cell(row, 2).value is not None
    }


def template_voltage_rows(ws) -> dict[str, int]:
    rows: dict[str, int] = {}
    for row in range(1, ws.max_row + 1):
        value = ws.cell(row, 2).value
        if not isinstance(value, str):
            continue
        match = re.search(r"_U=(\d+v)_", value, re.I)
        if match:
            rows[match.group(1).lower()] = row
    return rows


def template_frequency_columns(ws, block_row: int) -> dict[float, int]:
    columns: dict[float, int] = {}
    for column in range(1, ws.max_column + 1):
        value = ws.cell(block_row + 1, column).value
        if not isinstance(value, str):
            continue
        match = re.search(r"Частота\s+([0-9.,]+)\s*Гц", value, re.I)
        if match:
            columns[float(match.group(1).replace(",", "."))] = column
    return columns


def find_frequency_column(columns: dict[float, int], frequency: float) -> int:
    for available_frequency, column in columns.items():
        if math.isclose(available_frequency, frequency, rel_tol=1e-9, abs_tol=1e-9):
            return column
    raise RuntimeError(
        f"В Excel-шаблоне отсутствует подблок частоты {frequency:g} Гц."
    )


def validate_record_frequency(
    record: RwdRecord,
    identity: ExperimentIdentity,
    row_count: int,
    bindings: list[ChannelBinding],
) -> None:
    frequency_binding = next(
        (
            binding
            for binding in bindings
            if normalize_header(binding.header) == normalize_header("f in Hz")
        ),
        None,
    )
    if frequency_binding is None:
        raise RuntimeError("В эталонном ASCII отсутствует обязательная колонка 'f in Hz'.")
    actual_values = resolve_values(record, frequency_binding, row_count)
    if not all(value_matches(identity.frequency_sort_key, item.value) for item in actual_values):
        raise RuntimeError(
            f"Файл {record.path.name}: канал частоты не соответствует имени файла "
            f"({identity.frequency_sort_key:g} Гц). Для этой структуры нужен отдельный ASCII-эталон."
        )


def fill_template_measurement_blocks(
    ws,
    records: list[RwdRecord],
    ascii_record: AsciiRecord,
    bindings: list[ChannelBinding],
) -> bool:
    voltage_rows = template_voltage_rows(ws)
    if not voltage_rows:
        return False

    selected_bindings = display_bindings(bindings)
    row_count = len(ascii_record.rows)
    seen: set[tuple[str, float]] = set()
    for record in records:
        identity = parse_experiment_identity(record.path)
        voltage = identity.voltage.removeprefix("U=").lower()
        block_row = voltage_rows.get(voltage)
        if block_row is None:
            raise RuntimeError(
                f"В Excel-шаблоне отсутствует блок напряжения {identity.voltage!r}."
            )
        duplicate_key = (voltage, identity.frequency_sort_key)
        if duplicate_key in seen:
            raise RuntimeError(
                f"Для напряжения {identity.voltage!r} передано несколько файлов "
                f"частоты {identity.frequency_sort_key:g} Гц."
            )
        seen.add(duplicate_key)
        validate_record_frequency(record, identity, row_count, bindings)

        frequency_start_column = find_frequency_column(
            template_frequency_columns(ws, block_row),
            identity.frequency_sort_key,
        )
        rows = extract_signed_rows(record, row_count, selected_bindings)
        for row_offset, values in enumerate(rows, start=3):
            ws.cell(block_row + row_offset, frequency_start_column - 1, values[0])
            for column_offset, value in enumerate(values[1:]):
                ws.cell(block_row + row_offset, frequency_start_column + column_offset, value)
    return True


def append_measurement_blocks(
    ws,
    records: list[RwdRecord],
    ascii_record: AsciiRecord,
    bindings: list[ChannelBinding],
) -> None:
    if fill_template_measurement_blocks(ws, records, ascii_record, bindings):
        return

    selected_bindings = display_bindings(bindings)
    groups = group_records(records)
    existing_names = existing_block_names(ws)
    duplicates = [block_name for block_name, _ in groups if block_name.lower() in existing_names]
    if duplicates:
        raise RuntimeError(
            "В выбранном листе уже существуют блоки: "
            + ", ".join(duplicates)
            + ". Удалите старые блоки или используйте другой шаблон."
        )
    fill = PatternFill("solid", fgColor="1F4E78")
    row_count = len(ascii_record.rows)
    block_start_row = next_append_row(ws)

    for block_name, items in groups:
        ws.cell(block_start_row, 2, block_name)
        for frequency_index, (identity, record) in enumerate(items):
            frequency_start_column = 1 + frequency_index * 7
            value_start_column = frequency_start_column + 1
            ws.cell(block_start_row + 1, value_start_column, identity.frequency)
            for column_offset, (display_header, _) in enumerate(selected_bindings):
                ws.cell(block_start_row + 2, value_start_column + column_offset, display_header)

            rows = extract_signed_rows(record, row_count, selected_bindings)
            for row_offset, row in enumerate(rows, start=3):
                ws.cell(block_start_row + row_offset, frequency_start_column, row[0])
                for column_offset, value in enumerate(row[1:]):
                    ws.cell(block_start_row + row_offset, value_start_column + column_offset, value)

            for row in (block_start_row + 1, block_start_row + 2):
                for column in range(frequency_start_column, frequency_start_column + 7):
                    cell = ws.cell(row, column)
                    if cell.value is not None:
                        cell.font = Font(bold=True, color="FFFFFF")
                        cell.fill = fill
                        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            for column in range(frequency_start_column, frequency_start_column + 7):
                ws.column_dimensions[get_column_letter(column)].width = 18
        block_start_row += row_count + 5
    ws.freeze_panes = "B4"


def solve_linear_system(matrix: list[list[float]], vector: list[float]) -> list[float]:
    size = len(vector)
    augmented = [row[:] + [vector[index]] for index, row in enumerate(matrix)]
    for column in range(size):
        pivot = max(range(column, size), key=lambda row: abs(augmented[row][column]))
        if abs(augmented[pivot][column]) < 1e-12:
            raise RuntimeError("Не удалось рассчитать полиномиальную кривую: вырожденная система.")
        augmented[column], augmented[pivot] = augmented[pivot], augmented[column]
        divisor = augmented[column][column]
        augmented[column] = [value / divisor for value in augmented[column]]
        for row in range(size):
            if row == column:
                continue
            factor = augmented[row][column]
            augmented[row] = [
                current - factor * selected
                for current, selected in zip(augmented[row], augmented[column])
            ]
    return [augmented[row][-1] for row in range(size)]


def polynomial_coefficients(xs: list[float], ys: list[float], order: int) -> list[float]:
    powers = [sum(x ** exponent for x in xs) for exponent in range(order * 2 + 1)]
    matrix = [
        [powers[row + column] for column in range(order + 1)]
        for row in range(order + 1)
    ]
    vector = [
        sum(y * (x ** exponent) for x, y in zip(xs, ys))
        for exponent in range(order + 1)
    ]
    return solve_linear_system(matrix, vector)


def log_polynomial_curve(points: list[tuple[float, float]]) -> list[tuple[float, float]]:
    valid_points = [
        (x, y)
        for x, y in points
        if math.isfinite(x) and math.isfinite(y) and x > 0 and y > 0
    ]
    if len(valid_points) <= CHART_POLYNOMIAL_ORDER:
        return []
    log_xs = [math.log10(x) for x, _ in valid_points]
    log_ys = [math.log10(y) for _, y in valid_points]
    coefficients = polynomial_coefficients(log_xs, log_ys, CHART_POLYNOMIAL_ORDER)
    minimum = min(log_xs)
    maximum = max(log_xs)
    if math.isclose(minimum, maximum):
        return []
    return [
        (
            10 ** log_x,
            10 ** sum(coefficient * (log_x ** exponent) for exponent, coefficient in enumerate(coefficients)),
        )
        for log_x in (
            minimum + (maximum - minimum) * index / (CHART_CURVE_POINT_COUNT - 1)
            for index in range(CHART_CURVE_POINT_COUNT)
        )
    ]


def chart_input_points(
    ws,
    block_row: int,
    frequency_column: int,
    metric_index: int,
) -> list[tuple[float, float]]:
    points: list[tuple[float, float]] = []
    for row in range(block_row + 3, block_row + 23):
        gamma = ws.cell(row, frequency_column + 5).value
        if metric_index == 3:
            storage_modulus = ws.cell(row, frequency_column + 2).value
            loss_modulus = ws.cell(row, frequency_column + 3).value
            if not isinstance(storage_modulus, (int, float)) or storage_modulus == 0:
                continue
            y_value = loss_modulus / storage_modulus if isinstance(loss_modulus, (int, float)) else None
        else:
            y_value = ws.cell(
                row,
                frequency_column + CHART_METRIC_INPUT_OFFSETS[metric_index],
            ).value
        if isinstance(gamma, (int, float)) and isinstance(y_value, (int, float)):
            points.append((float(gamma), float(y_value)))
    return points


def chart_range_formula(sheet_name: str, column: int, row_start: int, row_end: int) -> str:
    letter = get_column_letter(column)
    return f"{sheet_name}!${letter}${row_start}:${letter}${row_end}"


def apply_marker_style(series, color: str, marker_symbol: str) -> None:
    series.graphicalProperties.line.noFill = True
    series.marker.symbol = marker_symbol
    series.marker.size = 5
    series.marker.graphicalProperties.solidFill = color
    series.marker.graphicalProperties.line.solidFill = color
    series.trendline = None


def apply_curve_style(series, color: str) -> None:
    series.graphicalProperties.line.noFill = False
    series.graphicalProperties.line.solidFill = color
    series.graphicalProperties.line.width = 19050
    series.marker.symbol = "none"
    series.trendline = None
    series.smooth = True


def clear_manual_chart_legend(ws, chart) -> None:
    start_row = chart.anchor._from.row + 34
    start_column = chart.anchor._from.col + 2
    for voltage_index in range(len(CHART_VOLTAGE_STYLES)):
        row = start_row + voltage_index % 3
        column = start_column + (voltage_index // 3) * 4
        ws.cell(row, column).value = None
        ws.cell(row, column + 1).value = None


def make_chart_layer_transparent(chart) -> None:
    transparent = GraphicalProperties(noFill=True, ln=LineProperties(noFill=True))
    chart.graphical_properties = deepcopy(transparent)
    chart.plot_area.spPr = deepcopy(transparent)


def refresh_chart_curves(workbook) -> None:
    if "freq_diagrams" not in workbook.sheetnames or "input_data" not in workbook.sheetnames:
        return
    diagram_ws = workbook["freq_diagrams"]
    if not diagram_ws._charts:
        return
    base_charts = [chart for chart in diagram_ws._charts if chart.title is not None]
    if len(base_charts) != 40:
        raise RuntimeError(
            f"На листе 'freq_diagrams' ожидалось 40 основных диаграмм, найдено: {len(base_charts)}."
        )
    diagram_ws._charts = base_charts

    curve_sheet_name = "chart_curves"
    if curve_sheet_name in workbook.sheetnames:
        del workbook[curve_sheet_name]
    curve_ws = workbook.create_sheet(curve_sheet_name)
    input_ws = workbook["input_data"]
    voltage_rows = template_voltage_rows(input_ws)
    if not voltage_rows:
        raise RuntimeError("Не удалось найти блоки напряжений в листе 'input_data'.")
    first_block_row = min(voltage_rows.values())
    frequencies = sorted(template_frequency_columns(input_ws, first_block_row).items())
    if len(frequencies) != 10:
        raise RuntimeError(
            f"В листе 'input_data' ожидалось 10 подблоков частот, найдено: {len(frequencies)}."
        )

    layered_charts = []
    for chart_index, chart in enumerate(base_charts):
        if not chart.ser:
            raise RuntimeError(f"Диаграмма {chart_index + 1}: отсутствует исходный ряд данных.")
        clear_manual_chart_legend(diagram_ws, chart)
        frequency_index = chart_index // 4
        metric_index = chart_index % 4
        frequency, frequency_column = frequencies[frequency_index]
        row_start = 5 + frequency_index * 32
        row_end = row_start + 19
        curve_row_start = 2 + chart_index * (CHART_CURVE_POINT_COUNT + 2)
        curve_row_end = curve_row_start + CHART_CURVE_POINT_COUNT - 1
        x_columns, y_columns = CHART_METRIC_COLUMNS[metric_index]
        raw_template = chart.ser[0]
        raw_series = []
        curve_series = []

        curve_ws.cell(curve_row_start - 1, 1, f"chart_{chart_index + 1}_{frequency:g}hz")
        for voltage_index, ((voltage, color, marker), x_column, y_column) in enumerate(
            zip(CHART_VOLTAGE_STYLES, x_columns, y_columns)
        ):
            raw = deepcopy(raw_template)
            raw.idx = voltage_index
            raw.order = voltage_index
            raw.tx = SeriesLabel(v=f"SCTNA=60p_Katal68=2p_PyrO=40p_{voltage}")
            raw.xVal = AxDataSource(
                numRef=NumRef(f=chart_range_formula("output_data", x_column, row_start, row_end))
            )
            raw.yVal = NumDataSource(
                numRef=NumRef(f=chart_range_formula("output_data", y_column, row_start, row_end))
            )
            apply_marker_style(raw, color, marker)
            raw_series.append(raw)

            curve_x_column = voltage_index * 2 + 1
            curve_y_column = curve_x_column + 1
            curve_ws.cell(curve_row_start - 1, curve_x_column, f"{voltage}_x")
            curve_ws.cell(curve_row_start - 1, curve_y_column, f"{voltage}_y")
            block_row = voltage_rows.get(voltage)
            curve = (
                log_polynomial_curve(
                    chart_input_points(input_ws, block_row, frequency_column, metric_index)
                )
                if block_row is not None
                else []
            )
            for row_offset, (x_value, y_value) in enumerate(curve):
                curve_ws.cell(curve_row_start + row_offset, curve_x_column, x_value)
                curve_ws.cell(curve_row_start + row_offset, curve_y_column, y_value)

            smooth = deepcopy(raw_template)
            smooth.idx = len(CHART_VOLTAGE_STYLES) + voltage_index
            smooth.order = len(CHART_VOLTAGE_STYLES) + voltage_index
            # Excel for macOS can ignore deleted legend entries for auxiliary
            # scatter series. Keep the title visually empty as a second guard.
            smooth.tx = SeriesLabel(v=" ")
            smooth.xVal = AxDataSource(
                numRef=NumRef(
                    f=chart_range_formula(
                        curve_sheet_name,
                        curve_x_column,
                        curve_row_start,
                        curve_row_end,
                    )
                )
            )
            smooth.yVal = NumDataSource(
                numRef=NumRef(
                    f=chart_range_formula(
                        curve_sheet_name,
                        curve_y_column,
                        curve_row_start,
                        curve_row_end,
                    )
                )
            )
            apply_curve_style(smooth, color)
            curve_series.append(smooth)

        chart.ser = raw_series
        chart.x_axis.delete = False
        chart.y_axis.delete = False
        chart.x_axis.crosses = None
        chart.y_axis.crosses = None
        chart.x_axis.crossesAt = chart.y_axis.scaling.min
        chart.y_axis.crossesAt = chart.x_axis.scaling.min
        chart.plot_area.layout = Layout(
            manualLayout=ManualLayout(x=0.12, y=0.08, w=0.78, h=0.68)
        )
        chart.legend = Legend(legendPos="b", overlay=False)
        make_chart_layer_transparent(chart)

        curve_chart = deepcopy(chart)
        curve_chart.ser = curve_series
        curve_chart.title = None
        curve_chart.legend = None
        curve_chart.x_axis.delete = True
        curve_chart.y_axis.delete = True
        curve_chart.x_axis.title = None
        curve_chart.y_axis.title = None
        curve_chart.x_axis.majorGridlines = None
        curve_chart.x_axis.minorGridlines = None
        curve_chart.y_axis.majorGridlines = None
        curve_chart.y_axis.minorGridlines = None
        curve_chart.anchor = deepcopy(chart.anchor)
        make_chart_layer_transparent(curve_chart)
        layered_charts.extend([curve_chart, chart])

    diagram_ws._charts = layered_charts
    curve_ws.sheet_state = "hidden"


def metadata_fields(record: RwdRecord) -> dict[str, Any]:
    return {
        "file_path": str(record.path),
        "size_bytes": record.size_bytes,
        "format": record.format_magic,
        "rheowin_version": record.rheowin_version,
        "date": record.date,
        "time": record.time,
        "instrument_model": record.instrument_model,
        "device": record.device,
        "geometry": record.geometry,
        "temperature_controller": record.temperature_controller,
        "gap": record.gap,
        "serial_numbers": record.serial_numbers,
        "driver_versions": record.driver_versions,
        "firmware_versions": record.firmware_versions,
        "job_file": record.job_file,
        "mode": record.mode,
        "numeric_blocks_extracted": len(record.numeric_blocks),
    }


def update_metadata_sheet(workbook, records: list[RwdRecord]) -> None:
    sheet_name = "instrument_metadata"
    ws = workbook[sheet_name] if sheet_name in workbook.sheetnames else workbook.create_sheet(sheet_name)
    existing: dict[str, dict[str, Any]] = {}
    if ws.max_row >= 1 and ws.max_column >= 2:
        fields = [ws.cell(row, 1).value for row in range(2, ws.max_row + 1)]
        for column in range(2, ws.max_column + 1):
            file_name = ws.cell(1, column).value
            if file_name:
                existing[str(file_name)] = {
                    str(field): ws.cell(row, column).value
                    for row, field in enumerate(fields, start=2)
                    if field is not None
                }
    for record in records:
        existing[record.path.name] = metadata_fields(record)

    field_names = list(metadata_fields(records[0]).keys())
    ws.delete_rows(1, ws.max_row)
    ws.delete_cols(1, ws.max_column)
    ws.cell(1, 1, "field")
    for column, file_name in enumerate(sorted(existing, key=str.lower), start=2):
        ws.cell(1, column, file_name)
        for row, field_name in enumerate(field_names, start=2):
            ws.cell(row, 1, field_name)
            ws.cell(row, column, existing[file_name].get(field_name, ""))
    style_table(ws)


def style_table(ws) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for column in range(1, ws.max_column + 1):
        width = max(len(str(ws.cell(row, column).value or "")) for row in range(1, min(ws.max_row, 500) + 1))
        ws.column_dimensions[get_column_letter(column)].width = min(max(width + 2, 12), 70)


def write_mapping_sheet(workbook, ascii_record: AsciiRecord, calibration_record: RwdRecord, bindings: list[ChannelBinding]) -> None:
    sheet_name = "parser_mapping"
    if sheet_name in workbook.sheetnames:
        del workbook[sheet_name]
    ws = workbook.create_sheet(sheet_name)
    rows = [
        ["calibration_ascii", str(ascii_record.path)],
        ["calibration_rwd", str(calibration_record.path)],
        [],
        ["header", "raw_offset_hex"],
    ]
    rows.extend([[binding.header, hex(binding.raw_offset)] for binding in bindings])
    for row in rows:
        ws.append(row)
    style_table(ws)
    ws.sheet_state = "hidden"


def write_diagnostics_sheet(workbook, records: list[RwdRecord]) -> None:
    sheet_name = "raw_values"
    if sheet_name in workbook.sheetnames:
        del workbook[sheet_name]
    ws = workbook.create_sheet(sheet_name)
    ws.append(
        [
            "file_name",
            "record_offset_hex",
            "code_1",
            "code_2",
            "dtype",
            "value_index",
            "value_offset_hex",
            "raw_hex_le",
            "float32_value",
        ]
    )
    for record in records:
        for block in record.numeric_blocks:
            for value_index, item in enumerate(block.values, start=1):
                ws.append(
                    [
                        record.path.name,
                        hex(block.offset),
                        block.code_1,
                        block.code_2,
                        block.dtype,
                        value_index,
                        hex(item.offset),
                        item.raw_hex_le,
                        excel_value(item.value),
                    ]
                )
    style_table(ws)


def output_path_for(template: Path, requested: Path | None) -> Path:
    if requested:
        return requested.expanduser().resolve()
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return template.with_name(f"{template.stem}_parsed_{timestamp}.xlsx")


def parse_args(argv: list[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Заполнить Excel-шаблон фактическими данными из HAAKE RheoWin .rwd после калибровки по ASCII."
    )
    parser.add_argument("template", type=Path, help="Первый файл: Excel-шаблон .xlsx.")
    parser.add_argument("ascii", type=Path, help="Второй файл: эталонный ASCII-экспорт RheoWin .txt/.asc/.csv.")
    parser.add_argument("rwd", nargs="+", type=Path, help="Один или несколько .rwd файлов либо папок с .rwd.")
    parser.add_argument("--sheet", required=True, help="Название листа Excel, в который добавляются группы данных.")
    parser.add_argument("-o", "--output", type=Path, help="Путь к новой итоговой книге .xlsx.")
    parser.add_argument("-r", "--recursive", action="store_true", help="Искать .rwd во вложенных папках.")
    parser.add_argument("--include-diagnostics", action="store_true", help="Добавить технический лист raw_values.")
    return parser.parse_args(argv)


def validate_inputs(args: argparse.Namespace) -> tuple[Path, Path, list[Path]]:
    template = args.template.expanduser().resolve()
    ascii_path = args.ascii.expanduser().resolve()
    if not template.is_file() or template.suffix.lower() != ".xlsx":
        raise RuntimeError("Первым аргументом должен быть существующий Excel-шаблон .xlsx.")
    if not ascii_path.is_file() or ascii_path.suffix.lower() not in ASCII_SUFFIXES:
        raise RuntimeError("Вторым аргументом должен быть существующий ASCII-файл .txt, .asc или .csv.")
    rwd_files = collect_rwd_files(args.rwd, recursive=args.recursive)
    if not rwd_files:
        raise RuntimeError("Не найдены .rwd файлы для обработки.")
    return template, ascii_path, rwd_files


def main(argv: list[str] | None = None) -> int:
    try:
        args = parse_args(argv or sys.argv[1:])
        template, ascii_path, rwd_files = validate_inputs(args)
        ascii_record = extract_ascii_record(ascii_path)
        records = [extract_rwd_record(path) for path in rwd_files]
        calibration_record, bindings = calibrate(ascii_record, records)

        workbook = load_workbook(template)
        if args.sheet not in workbook.sheetnames:
            raise RuntimeError(
                f"В шаблоне нет листа {args.sheet!r}. Доступные листы: {', '.join(workbook.sheetnames)}"
            )
        append_measurement_blocks(workbook[args.sheet], records, ascii_record, bindings)
        refresh_chart_curves(workbook)
        update_metadata_sheet(workbook, records)
        write_mapping_sheet(workbook, ascii_record, calibration_record, bindings)
        if args.include_diagnostics:
            write_diagnostics_sheet(workbook, records)

        output = output_path_for(template, args.output)
        workbook.calculation.fullCalcOnLoad = True
        workbook.calculation.forceFullCalc = True
        workbook.calculation.calcMode = "auto"
        workbook.save(output)
        print(f"Калибровочный .rwd: {calibration_record.path.name}")
        print(f"Сопоставлено колонок: {len(bindings)}")
        print(f"Обработано .rwd файлов: {len(records)}")
        print(f"Создан файл: {output}")
        return 0
    except RuntimeError as error:
        print(f"Ошибка: {error}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
